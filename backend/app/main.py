import csv
import io
import logging
import os
import threading
from datetime import date
from pathlib import Path
from typing import List
from dotenv import load_dotenv
from fastapi import Depends, FastAPI, File, HTTPException, Request, UploadFile, status
from fastapi.responses import StreamingResponse
from fastapi.middleware.cors import CORSMiddleware
from sqlalchemy import text
from sqlalchemy.orm import Session
from . import crud, models, schemas, auth, schedule
from .database import engine
from .dependencies import get_current_active_user, get_current_superuser, get_db

load_dotenv(Path(__file__).resolve().parent.parent / ".env")
load_dotenv(Path(__file__).resolve().parent.parent / ".env.local", override=True)

logger = logging.getLogger(__name__)

ADMIN_EMAIL = "toshmirzayevinomjon@gmail.com"

BOT_ENABLED = False
if os.environ.get("TELEGRAM_BOT_TOKEN"):
    try:
        from . import telegram_bot

        BOT_ENABLED = True
    except Exception as e:
        logger.warning(f"Telegram bot module could not be loaded: {e}")

models.Base.metadata.create_all(bind=engine)
try:
    with engine.connect() as conn:
        conn.execute(
            text(
                "ALTER TABLE groups ADD COLUMN is_archived BOOLEAN NOT NULL DEFAULT false"
            )
        )
        conn.commit()
except Exception:
    pass
try:
    with engine.connect() as conn:
        conn.execute(
            text("ALTER TABLE groups ADD COLUMN monthly_fee INTEGER NOT NULL DEFAULT 0")
        )
        conn.commit()
except Exception:
    pass
try:
    with engine.connect() as conn:
        conn.execute(text("ALTER TABLE users ADD COLUMN telegram_chat_id VARCHAR(255)"))
        conn.execute(
            text(
                "ALTER TABLE users ADD COLUMN language VARCHAR(10) NOT NULL DEFAULT 'uz'"
            )
        )
        conn.commit()
except Exception:
    pass
try:
    with engine.connect() as conn:
        conn.execute(
            text("ALTER TABLE users ADD COLUMN monthly_fee INTEGER NOT NULL DEFAULT 0")
        )
        conn.commit()
except Exception:
    pass
try:
    with engine.connect() as conn:
        conn.execute(text("ALTER TABLE users ADD COLUMN card_number VARCHAR(50)"))
        conn.commit()
except Exception:
    pass
try:
    models.LessonTopic.__table__.create(bind=engine, checkfirst=True)
except Exception:
    pass
try:
    models.Payment.__table__.create(bind=engine, checkfirst=True)
except Exception:
    pass
try:
    models.ExceptionDay.__table__.create(bind=engine, checkfirst=True)
except Exception:
    pass
try:
    models.ParentLink.__table__.create(bind=engine, checkfirst=True)
except Exception:
    pass
try:
    with engine.connect() as conn:
        conn.execute(text("ALTER TABLE students ADD COLUMN parent_name VARCHAR(255)"))
        conn.execute(text("ALTER TABLE students ADD COLUMN parent_phone VARCHAR(50)"))
        conn.commit()
except Exception:
    pass
try:
    with engine.connect() as conn:
        conn.execute(text("ALTER TABLE students ADD COLUMN birth_date DATE"))
        conn.commit()
except Exception:
    pass
try:
    with engine.connect() as conn:
        conn.execute(text("ALTER TABLE lesson_topics ADD COLUMN homework TEXT"))
        conn.commit()
except Exception:
    pass
try:
    with engine.connect() as conn:
        conn.execute(
            text(
                "ALTER TABLE parent_links ADD COLUMN language VARCHAR(10) NOT NULL DEFAULT 'uz'"
            )
        )
        conn.commit()
except Exception:
    pass
try:
    models.PaymentReceipt.__table__.create(bind=engine, checkfirst=True)
except Exception:
    pass
try:
    models.Grade.__table__.create(bind=engine, checkfirst=True)
except Exception:
    pass
try:
    models.LoginLog.__table__.create(bind=engine, checkfirst=True)
except Exception:
    pass

app = FastAPI(title="Kundalik LMS Attendance API")
app.add_middleware(
    CORSMiddleware,
    allow_origins=[
        "https://kundalik.up.railway.app",
        "https://kundalik-back.up.railway.app",
        "http://127.0.0.1:5174",
        "http://localhost:5174",
    ],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


def validate_year_month(year: int, month: int) -> None:
    if year < 2020 or year > 2100:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Year must be between 2020 and 2100",
        )
    if month < 1 or month > 12:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Month must be between 1 and 12",
        )


@app.on_event("startup")
def start_telegram_bot():
    if BOT_ENABLED:

        def bot_runner():
            try:
                telegram_bot.run_bot()
            except Exception as e:
                logger.warning(
                    f"Telegram bot not available (expected if no internet): {e}"
                )

        t = threading.Thread(target=bot_runner, daemon=True)
        t.start()
        logger.info("Telegram bot started")


@app.get("/")
def read_root():
    return {"status": "ok", "service": "Kundalik LMS Attendance API"}


@app.get("/health")
def health_check():
    return {"status": "ok"}


@app.post("/auth/register", response_model=schemas.UserProfile)
def register(user_in: schemas.UserCreate, db: Session = Depends(get_db)):
    existing = crud.get_user_by_email(db, user_in.email)
    if existing:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST, detail="Email already registered"
        )
    if user_in.username:
        existing_username = crud.get_user_by_username(db, user_in.username)
        if existing_username:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST, detail="Username already taken"
            )
    user = crud.create_user(
        db,
        user_in,
        is_superuser=(user_in.email.strip().lower() == ADMIN_EMAIL),
    )
    return user


@app.post("/auth/login", response_model=schemas.Token)
def login(
    form_data: schemas.UserLogin, request: Request, db: Session = Depends(get_db)
):
    user = crud.get_user_by_email(db, form_data.login)
    if not user:
        user = crud.get_user_by_username(db, form_data.login)
    success = bool(
        user and auth.verify_password(form_data.password, user.password_hash)
    )
    ip = request.client.host if request.client else None
    ua = request.headers.get("user-agent", "")[:500]
    db.add(
        models.LoginLog(
            user_id=user.id if user else None,
            email=form_data.login,
            ip=ip,
            user_agent=ua,
            success=success,
        )
    )
    db.commit()
    if not success:
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED, detail="Invalid credentials"
        )
    if not user.is_active:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Hisob bloklangan. Admin bilan bog'laning.",
        )
    token = auth.create_access_token(subject=user.email)
    return {"access_token": token, "token_type": "bearer"}


@app.get("/auth/me", response_model=schemas.UserProfile)
def get_profile(current_user: models.User = Depends(get_current_active_user)):
    return current_user


@app.put("/auth/me", response_model=schemas.UserProfile)
def update_profile(
    upd: schemas.UserUpdate,
    current_user: models.User = Depends(get_current_active_user),
    db: Session = Depends(get_db),
):
    if upd.full_name is not None:
        current_user.full_name = upd.full_name.strip()
    if upd.telegram_chat_id is not None:
        current_user.telegram_chat_id = upd.telegram_chat_id.strip() or None
    if upd.language is not None:
        current_user.language = upd.language
    if upd.monthly_fee is not None:
        current_user.monthly_fee = upd.monthly_fee
    db.commit()
    db.refresh(current_user)
    return current_user


@app.post("/auth/link-code")
def generate_link_code(
    current_user: models.User = Depends(get_current_active_user),
):
    if not BOT_ENABLED:
        raise HTTPException(status_code=400, detail="Telegram bot is not configured")
    code = telegram_bot.generate_link_code(current_user.id)
    return {"code": code, "expires_in": 5}


@app.post("/auth/change-password")
def change_password(
    pw: schemas.PasswordChange,
    current_user: models.User = Depends(get_current_active_user),
    db: Session = Depends(get_db),
):
    if not auth.verify_password(pw.current_password, current_user.password_hash):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST, detail="Joriy parol noto'g'ri"
        )
    current_user.password_hash = auth.hash_password(pw.new_password)
    db.commit()
    return {"message": "Parol muvaffaqiyatli o'zgartirildi"}


@app.get("/groups", response_model=List[schemas.GroupOut])
def list_groups(
    archived: bool = False,
    current_user: models.User = Depends(get_current_active_user),
    db: Session = Depends(get_db),
):
    return crud.get_groups_for_user(db, current_user.id, include_archived=archived)


@app.post("/groups", response_model=schemas.GroupOut)
def create_group(
    group_in: schemas.GroupCreate,
    current_user: models.User = Depends(get_current_active_user),
    db: Session = Depends(get_db),
):
    return crud.create_group(db, current_user.id, group_in)


@app.put("/groups/{group_id}", response_model=schemas.GroupOut)
def update_group(
    group_id: int,
    group_in: schemas.GroupUpdate,
    current_user: models.User = Depends(get_current_active_user),
    db: Session = Depends(get_db),
):
    group = crud.get_group_by_id(db, group_id, current_user.id)
    if not group:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND, detail="Group not found"
        )
    return crud.update_group(db, group, group_in)


@app.delete("/groups/{group_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_group(
    group_id: int,
    current_user: models.User = Depends(get_current_active_user),
    db: Session = Depends(get_db),
):
    group = crud.get_group_by_id(db, group_id, current_user.id)
    if not group:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND, detail="Group not found"
        )
    crud.delete_group(db, group)


@app.patch("/groups/{group_id}/archive", response_model=schemas.GroupOut)
def archive_group(
    group_id: int,
    current_user: models.User = Depends(get_current_active_user),
    db: Session = Depends(get_db),
):
    group = crud.get_group_by_id(db, group_id, current_user.id)
    if not group:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND, detail="Group not found"
        )
    group.is_archived = not group.is_archived
    db.commit()
    db.refresh(group)
    return group


@app.get("/groups/{group_id}/students", response_model=List[schemas.StudentOut])
def list_students(
    group_id: int,
    current_user: models.User = Depends(get_current_active_user),
    db: Session = Depends(get_db),
):
    group = crud.get_group_by_id(db, group_id, current_user.id)
    if not group:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND, detail="Group not found"
        )
    return crud.get_students_by_group(db, group_id)


@app.post("/groups/{group_id}/students", response_model=schemas.StudentOut)
def create_student(
    group_id: int,
    student_in: schemas.StudentCreate,
    current_user: models.User = Depends(get_current_active_user),
    db: Session = Depends(get_db),
):
    group = crud.get_group_by_id(db, group_id, current_user.id)
    if not group:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND, detail="Group not found"
        )
    return crud.create_student(db, group_id, student_in)


@app.post("/groups/{group_id}/students/import", response_model=List[schemas.StudentOut])
def import_students_csv(
    group_id: int,
    file: UploadFile = File(...),
    current_user: models.User = Depends(get_current_active_user),
    db: Session = Depends(get_db),
):
    group = crud.get_group_by_id(db, group_id, current_user.id)
    if not group:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND, detail="Group not found"
        )
    if not file.filename or not file.filename.endswith(".csv"):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST, detail="Faqat CSV fayl yuklang"
        )
    content = file.file.read().decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(content))
    if "full_name" not in (reader.fieldnames or []):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="CSV da 'full_name' ustuni bo'lishi kerak",
        )
    students_data = []
    for row in reader:
        name = row.get("full_name", "").strip()
        if name:
            students_data.append(
                {
                    "full_name": name,
                    "phone": row.get("phone", "").strip() or None,
                    "parent_name": row.get("parent_name", "").strip() or None,
                    "parent_phone": row.get("parent_phone", "").strip() or None,
                    "birth_date": row.get("birth_date", "").strip() or None,
                }
            )
    if not students_data:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="CSV dan hech qanday o'quvchi topilmadi",
        )
    return crud.bulk_create_students(db, group_id, students_data)


@app.get("/groups/{group_id}/students/template")
def student_csv_template(
    group_id: int,
    current_user: models.User = Depends(get_current_active_user),
    db: Session = Depends(get_db),
):
    group = crud.get_group_by_id(db, group_id, current_user.id)
    if not group:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND, detail="Group not found"
        )
    buffer = io.StringIO()
    writer = csv.writer(buffer)
    writer.writerow(["full_name", "phone", "parent_name", "parent_phone", "birth_date"])
    writer.writerow(
        [
            "Aliyev Alisher",
            "+998901234567",
            "Aliyev Akmal",
            "+998901234568",
            "2012-03-15",
        ]
    )
    writer.writerow(["Karimova Malika", "+998937654321", "", "", "2013-07-22"])
    buffer.seek(0)
    return StreamingResponse(
        iter([buffer.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": 'attachment; filename="students_template.csv"'},
    )


@app.get("/search")
def global_search(
    q: str,
    current_user: models.User = Depends(get_current_active_user),
    db: Session = Depends(get_db),
):
    from sqlalchemy import or_

    query = (q or "").strip()
    if not query:
        return {"groups": [], "students": []}
    like = f"%{query}%"
    groups = (
        db.query(models.Group)
        .filter(
            models.Group.user_id == current_user.id,
            models.Group.is_archived.is_(False),
            models.Group.name.ilike(like),
        )
        .order_by(models.Group.name.asc())
        .limit(10)
        .all()
    )
    rows = (
        db.query(models.Student, models.Group.name)
        .join(models.Group, models.Student.group_id == models.Group.id)
        .filter(
            models.Group.user_id == current_user.id,
            models.Group.is_archived.is_(False),
            models.Student.status == "active",
            or_(
                models.Student.full_name.ilike(like),
                models.Student.phone.ilike(like),
            ),
        )
        .order_by(models.Student.full_name.asc())
        .limit(10)
        .all()
    )
    return {
        "groups": [{"id": g.id, "name": g.name} for g in groups],
        "students": [
            {
                "id": s.id,
                "full_name": s.full_name,
                "phone": s.phone,
                "group_id": s.group_id,
                "group_name": group_name,
            }
            for s, group_name in rows
        ],
    }


@app.put("/groups/{group_id}/students/{student_id}", response_model=schemas.StudentOut)
def update_student(
    group_id: int,
    student_id: int,
    student_in: schemas.StudentUpdate,
    current_user: models.User = Depends(get_current_active_user),
    db: Session = Depends(get_db),
):
    group = crud.get_group_by_id(db, group_id, current_user.id)
    if not group:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND, detail="Group not found"
        )
    student = crud.get_student_by_id(db, student_id, group_id)
    if not student:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND, detail="Student not found"
        )
    return crud.update_student(db, student, student_in)


@app.delete(
    "/groups/{group_id}/students/{student_id}", status_code=status.HTTP_204_NO_CONTENT
)
def delete_student(
    group_id: int,
    student_id: int,
    current_user: models.User = Depends(get_current_active_user),
    db: Session = Depends(get_db),
):
    group = crud.get_group_by_id(db, group_id, current_user.id)
    if not group:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND, detail="Group not found"
        )
    student = crud.get_student_by_id(db, student_id, group_id)
    if not student:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND, detail="Student not found"
        )
    crud.delete_student(db, student)


@app.post("/attendance/batch", response_model=List[schemas.AttendanceOut])
def batch_attendance(
    batch: schemas.AttendanceBatchRequest,
    current_user: models.User = Depends(get_current_active_user),
    db: Session = Depends(get_db),
):
    group = crud.get_group_by_id(db, batch.group_id, current_user.id)
    if not group:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND, detail="Group not found"
        )
    student_ids = {student.id for student in group.students}
    for record in batch.records:
        if record.student_id not in student_ids:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"Student {record.student_id} does not belong to group",
            )
    saved = crud.batch_save_attendance(db, batch.group_id, batch.date, batch.records)
    if BOT_ENABLED and saved:
        try:
            for record in batch.records:
                if record.status and record.status.value == "absent":
                    student = (
                        db.query(models.Student)
                        .filter(models.Student.id == record.student_id)
                        .first()
                    )
                    if student:
                        telegram_bot.send_absent_notification(
                            group_name=group.name,
                            student_name=student.full_name,
                            student_id=student.id,
                            date_str=batch.date.isoformat(),
                            teacher_chat_id=current_user.telegram_chat_id,
                        )
        except Exception as e:
            logger.error(f"Failed to send absent notification: {e}")
    return saved


@app.get(
    "/groups/{group_id}/exception-days", response_model=List[schemas.ExceptionDayOut]
)
def list_exception_days(
    group_id: int,
    year: int,
    month: int,
    current_user: models.User = Depends(get_current_active_user),
    db: Session = Depends(get_db),
):
    validate_year_month(year, month)
    group = crud.get_group_by_id(db, group_id, current_user.id)
    if not group:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND, detail="Group not found"
        )
    return crud.get_exception_days_for_month(db, group_id, year, month)


@app.post("/groups/{group_id}/exception-days", response_model=schemas.ExceptionDayOut)
def create_exception_day(
    group_id: int,
    day_in: schemas.ExceptionDayCreate,
    current_user: models.User = Depends(get_current_active_user),
    db: Session = Depends(get_db),
):
    group = crud.get_group_by_id(db, group_id, current_user.id)
    if not group:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND, detail="Group not found"
        )
    existing = (
        db.query(models.ExceptionDay)
        .filter(
            models.ExceptionDay.group_id == group_id,
            models.ExceptionDay.date == day_in.date,
        )
        .first()
    )
    if existing:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Bu kun allaqachon istisno sifatida belgilangan",
        )
    day = models.ExceptionDay(group_id=group_id, date=day_in.date, reason=day_in.reason)
    db.add(day)
    db.commit()
    db.refresh(day)
    return day


@app.delete(
    "/groups/{group_id}/exception-days/{day_id}", status_code=status.HTTP_204_NO_CONTENT
)
def delete_exception_day(
    group_id: int,
    day_id: int,
    current_user: models.User = Depends(get_current_active_user),
    db: Session = Depends(get_db),
):
    group = crud.get_group_by_id(db, group_id, current_user.id)
    if not group:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND, detail="Group not found"
        )
    day = (
        db.query(models.ExceptionDay)
        .filter(
            models.ExceptionDay.id == day_id, models.ExceptionDay.group_id == group_id
        )
        .first()
    )
    if not day:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND, detail="Exception day not found"
        )
    db.delete(day)
    db.commit()


@app.get("/groups/{group_id}/export/pdf")
def export_pdf(
    group_id: int,
    year: int,
    month: int,
    current_user: models.User = Depends(get_current_active_user),
    db: Session = Depends(get_db),
):
    from calendar import monthrange
    from fpdf import FPDF
    import io

    validate_year_month(year, month)
    group = crud.get_group_by_id(db, group_id, current_user.id)
    if not group:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND, detail="Group not found"
        )

    students = crud.get_students_by_group(db, group_id)
    attendances = crud.get_attendance_for_group_month(db, group_id, year, month)
    dates = schedule.get_group_dates_for_month(group.days_of_week, year, month)
    exc_days = {
        e.date for e in crud.get_exception_days_for_month(db, group_id, year, month)
    }
    dates = [d for d in dates if d not in exc_days]

    attend_by_student: dict = {}
    for a in attendances:
        attend_by_student.setdefault(a.student_id, {})[a.date.isoformat()] = a.status

    pdf = FPDF(orientation="L", unit="mm", format="A4")
    pdf.add_page()
    pdf.set_font("Helvetica", size=7)

    pdf.set_fill_color(5, 150, 105)
    pdf.set_text_color(255, 255, 255)
    pdf.cell(50, 7, "O'quvchi", border=1, fill=True)
    for d in dates:
        pdf.cell(12, 7, d.strftime("%d.%m"), border=1, fill=True, align="C")
    pdf.cell(15, 7, "Kelgan", border=1, fill=True, align="C")
    pdf.cell(12, 7, "Foiz", border=1, fill=True, align="C")
    pdf.ln()

    pdf.set_text_color(30, 41, 59)
    for s in students:
        present = 0
        pdf.cell(50, 6, s.full_name[:25], border=1)
        for d in dates:
            st = attend_by_student.get(s.id, {}).get(d.isoformat())
            val = {"present": "K", "absent": "Y", "excused": "S"}.get(st, "")
            if st == "present":
                pdf.set_fill_color(209, 250, 229)
            elif st == "absent":
                pdf.set_fill_color(254, 226, 226)
            elif st == "excused":
                pdf.set_fill_color(254, 243, 199)
            else:
                pdf.set_fill_color(255, 255, 255)
            pdf.cell(12, 6, val, border=1, fill=True, align="C")
            if st == "present":
                present += 1
        pct = round(present / len(dates) * 100) if dates else 0
        pdf.cell(15, 6, str(present), border=1, align="C")
        pdf.cell(12, 6, f"{pct}%", border=1, align="C")
        pdf.ln()

    buf = io.BytesIO(pdf.output())
    return StreamingResponse(
        buf,
        media_type="application/pdf",
        headers={
            "Content-Disposition": f"attachment; filename={group.name}_{month}_{year}.pdf"
        },
    )


@app.post(
    "/groups/{group_id}/students/batch-status", response_model=List[schemas.StudentOut]
)
def batch_student_status(
    group_id: int,
    body: dict,
    current_user: models.User = Depends(get_current_active_user),
    db: Session = Depends(get_db),
):
    group = crud.get_group_by_id(db, group_id, current_user.id)
    if not group:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND, detail="Group not found"
        )
    ids = body.get("student_ids", [])
    status = body.get("status", "active")
    if not ids:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST, detail="No students selected"
        )
    return crud.batch_update_student_status(db, group_id, ids, status)


@app.get("/groups/{group_id}/payments/summary")
def payment_summary(
    group_id: int,
    year: int,
    month: int,
    current_user: models.User = Depends(get_current_active_user),
    db: Session = Depends(get_db),
):
    validate_year_month(year, month)
    group = crud.get_group_by_id(db, group_id, current_user.id)
    if not group:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND, detail="Group not found"
        )
    from calendar import monthrange
    from sqlalchemy import func

    result = (
        db.query(
            func.count(models.Payment.id),
            func.coalesce(func.sum(models.Payment.amount), 0),
        )
        .filter(
            models.Payment.group_id == group_id,
            models.Payment.date >= date(year, month, 1),
            models.Payment.date <= date(year, month, monthrange(year, month)[1]),
        )
        .first()
    )
    payments = (
        db.query(models.Payment)
        .filter(
            models.Payment.group_id == group_id,
            models.Payment.date >= date(year, month, 1),
            models.Payment.date <= date(year, month, monthrange(year, month)[1]),
        )
        .order_by(models.Payment.date.asc())
        .all()
    )
    return {
        "total_income": result[1] or 0,
        "student_count": len(set(p.student_id for p in payments)),
        "payments": payments,
    }


@app.get("/stats", response_model=schemas.StatsOut)
def get_stats(
    current_user: models.User = Depends(get_current_active_user),
    db: Session = Depends(get_db),
):
    today = date.today()
    months = []
    y, m = today.year, today.month
    for _ in range(6):
        months.append(f"{y:04d}-{m:02d}")
        m -= 1
        if m == 0:
            m = 12
            y -= 1
    months.reverse()
    window_start = date(int(months[0][:4]), int(months[0][5:7]), 1)
    month_start = today.replace(day=1)

    group_ids = [
        gid
        for (gid,) in db.query(models.Group.id).filter(
            models.Group.user_id == current_user.id
        )
    ]
    if not group_ids:
        return {
            "total_students": 0,
            "month_payments": 0,
            "month_attendance_rate": 0.0,
            "payments_by_month": [{"month": m, "total": 0} for m in months],
            "attendance_by_month": [{"month": m, "rate": 0.0} for m in months],
            "groups": [],
            "students": [],
        }

    payments = (
        db.query(models.Payment)
        .filter(
            models.Payment.group_id.in_(group_ids),
            models.Payment.date >= window_start,
        )
        .all()
    )
    pays_map = {m: 0 for m in months}
    month_payments = 0
    for p in payments:
        key = f"{p.date.year:04d}-{p.date.month:02d}"
        if key in pays_map:
            pays_map[key] += p.amount
            if p.date >= month_start:
                month_payments += p.amount
    payments_by_month = [{"month": m, "total": pays_map[m]} for m in months]

    attendance = (
        db.query(models.Attendance)
        .filter(
            models.Attendance.group_id.in_(group_ids),
            models.Attendance.date >= window_start,
        )
        .all()
    )
    att_map = {m: [0, 0] for m in months}
    month_att_total = 0
    month_att_present = 0
    for a in attendance:
        key = f"{a.date.year:04d}-{a.date.month:02d}"
        if key not in att_map:
            continue
        if a.date >= month_start:
            month_att_total += 1
            if a.status == "present":
                month_att_present += 1
        if a.status == "present":
            att_map[key][0] += 1
        att_map[key][1] += 1

    attendance_by_month = []
    for m in months:
        present, total = att_map[m]
        attendance_by_month.append(
            {
                "month": m,
                "rate": round(present / total * 100, 1) if total else 0.0,
            }
        )

    groups = (
        db.query(models.Group)
        .filter(
            models.Group.user_id == current_user.id,
            models.Group.is_archived.is_(False),
        )
        .order_by(models.Group.name.asc())
        .all()
    )

    all_attendance = (
        db.query(models.Attendance)
        .filter(models.Attendance.group_id.in_(group_ids))
        .all()
    )
    student_att = {}
    group_att = {}
    for a in all_attendance:
        bucket = student_att.setdefault(a.student_id, [0, 0, 0])
        if a.status == "present":
            bucket[0] += 1
        elif a.status == "absent":
            bucket[1] += 1
        elif a.status == "excused":
            bucket[2] += 1
        gb = group_att.setdefault(a.group_id, [0, 0])
        if a.status == "present":
            gb[0] += 1
        gb[1] += 1

    students = (
        db.query(models.Student)
        .filter(
            models.Student.group_id.in_([g.id for g in groups]),
            models.Student.status == "active",
        )
        .all()
    )
    student_stats = []
    for s in students:
        p, ab, ex = student_att.get(s.id, [0, 0, 0])
        total = p + ab + ex
        student_stats.append(
            {
                "id": s.id,
                "group_id": s.group_id,
                "full_name": s.full_name,
                "present": p,
                "absent": ab,
                "excused": ex,
                "attendance_rate": round(p / total * 100, 1) if total else 0.0,
            }
        )

    group_stats = []
    for g in groups:
        present, total = group_att.get(g.id, [0, 0])
        group_stats.append(
            {
                "id": g.id,
                "name": g.name,
                "student_count": sum(1 for s in students if s.group_id == g.id),
                "attendance_rate": round(present / total * 100, 1) if total else 0.0,
            }
        )

    month_present, month_total = att_map[months[-1]]
    return {
        "total_students": len(student_stats),
        "month_payments": month_payments,
        "month_attendance_rate": round(month_present / month_total * 100, 1)
        if month_total
        else 0.0,
        "payments_by_month": payments_by_month,
        "attendance_by_month": attendance_by_month,
        "groups": group_stats,
        "students": student_stats,
    }


@app.get("/groups/{group_id}/attendance", response_model=List[schemas.AttendanceOut])
def get_attendance(
    group_id: int,
    year: int,
    month: int,
    current_user: models.User = Depends(get_current_active_user),
    db: Session = Depends(get_db),
):
    validate_year_month(year, month)
    group = crud.get_group_by_id(db, group_id, current_user.id)
    if not group:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND, detail="Group not found"
        )
    return crud.get_attendance_for_group_month(db, group_id, year, month)


@app.get("/groups/{group_id}/export/excel")
def export_excel(
    group_id: int,
    year: int,
    month: int,
    current_user: models.User = Depends(get_current_active_user),
    db: Session = Depends(get_db),
):
    from calendar import monthrange
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    import io

    validate_year_month(year, month)
    group = crud.get_group_by_id(db, group_id, current_user.id)
    if not group:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND, detail="Group not found"
        )

    students = crud.get_students_by_group(db, group_id)
    attendances = crud.get_attendance_for_group_month(db, group_id, year, month)
    schedule = (
        db.query(models.LessonTopic)
        .filter(
            models.LessonTopic.group_id == group_id,
            models.LessonTopic.date >= date(year, month, 1),
            models.LessonTopic.date <= date(year, month, monthrange(year, month)[1]),
        )
        .all()
    )
    topics_map = {t.date.isoformat(): t.topic for t in schedule}
    from .schedule import get_group_dates_for_month

    dates = get_group_dates_for_month(group.days_of_week, year, month)

    wb = Workbook()
    ws = wb.active
    ws.title = f"{group.name} {month}.{year}"

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(
        start_color="059669", end_color="0D9488", fill_type="solid"
    )
    thin = Border(
        left=Side(style="thin", color="D1D5DB"),
        right=Side(style="thin", color="D1D5DB"),
        top=Side(style="thin", color="D1D5DB"),
        bottom=Side(style="thin", color="D1D5DB"),
    )

    ws.cell(1, 1, "O'quvchi").font = header_font
    ws.cell(1, 1).fill = header_fill
    ws.cell(1, 1).border = thin
    for i, d in enumerate(dates):
        cell = ws.cell(1, i + 2, d.strftime("%d.%m"))
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin
        cell.alignment = Alignment(horizontal="center")
    total_cell = ws.cell(1, len(dates) + 2, "Kelgan")
    total_cell.font = header_font
    total_cell.fill = header_fill
    total_cell.border = thin
    pct_cell = ws.cell(1, len(dates) + 3, "Foiz")
    pct_cell.font = header_font
    pct_cell.fill = header_fill
    pct_cell.border = thin
    topic_cell = ws.cell(1, len(dates) + 4, "Mavzu")
    topic_cell.font = header_font
    topic_cell.fill = header_fill
    topic_cell.border = thin

    attend_by_student: dict = {}
    for a in attendances:
        attend_by_student.setdefault(a.student_id, {})[a.date.isoformat()] = a.status

    green_fill = PatternFill(
        start_color="D1FAE5", end_color="D1FAE5", fill_type="solid"
    )
    red_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
    amber_fill = PatternFill(
        start_color="FEF3C7", end_color="FEF3C7", fill_type="solid"
    )

    for row_idx, student in enumerate(students, 2):
        ws.cell(row_idx, 1, student.full_name).border = thin
        present_count = 0
        for col_idx, d in enumerate(dates):
            ds = d.isoformat()
            status = attend_by_student.get(student.id, {}).get(ds)
            cell = ws.cell(row_idx, col_idx + 2)
            if status == "present":
                cell.value = "K"
                cell.fill = green_fill
                present_count += 1
            elif status == "absent":
                cell.value = "Y"
                cell.fill = red_fill
            elif status == "excused":
                cell.value = "S"
                cell.fill = amber_fill
            else:
                cell.value = ""
            cell.border = thin
            cell.alignment = Alignment(horizontal="center")
        ws.cell(row_idx, len(dates) + 2, present_count).border = thin
        pct = round(present_count / len(dates) * 100, 1) if dates else 0
        ws.cell(row_idx, len(dates) + 3, f"{pct}%").border = thin
        ws.cell(row_idx, len(dates) + 4, "").border = thin

    ws.column_dimensions["A"].width = 35
    for i in range(len(dates)):
        ws.column_dimensions[chr(66 + i)].width = 8

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename={group.name}_{month}_{year}.xlsx"
        },
    )


@app.get("/groups/{group_id}/topics", response_model=List[schemas.LessonTopicOut])
def list_topics(
    group_id: int,
    year: int,
    month: int,
    current_user: models.User = Depends(get_current_active_user),
    db: Session = Depends(get_db),
):
    validate_year_month(year, month)
    group = crud.get_group_by_id(db, group_id, current_user.id)
    if not group:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND, detail="Group not found"
        )
    return crud.get_lesson_topics(db, group_id, year, month)


@app.put("/groups/{group_id}/topics", response_model=schemas.LessonTopicOut)
def upsert_topic(
    group_id: int,
    topic_in: schemas.LessonTopicCreate,
    current_user: models.User = Depends(get_current_active_user),
    db: Session = Depends(get_db),
):
    group = crud.get_group_by_id(db, group_id, current_user.id)
    if not group:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND, detail="Group not found"
        )
    return crud.upsert_lesson_topic(db, group_id, topic_in)


@app.get("/groups/{group_id}/grades", response_model=List[schemas.GradeOut])
def list_grades(
    group_id: int,
    year: int,
    month: int,
    current_user: models.User = Depends(get_current_active_user),
    db: Session = Depends(get_db),
):
    validate_year_month(year, month)
    group = crud.get_group_by_id(db, group_id, current_user.id)
    if not group:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND, detail="Group not found"
        )
    return crud.get_grades_for_month(db, group_id, year, month)


@app.put("/groups/{group_id}/grades", response_model=schemas.GradeOut)
def upsert_grade(
    group_id: int,
    grade_in: schemas.GradeCreate,
    current_user: models.User = Depends(get_current_active_user),
    db: Session = Depends(get_db),
):
    group = crud.get_group_by_id(db, group_id, current_user.id)
    if not group:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND, detail="Group not found"
        )
    student = crud.get_student_by_id(db, grade_in.student_id, group_id)
    if not student:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND, detail="Student not found"
        )
    return crud.upsert_grade(db, group_id, grade_in)


@app.delete(
    "/groups/{group_id}/grades/{student_id}", status_code=status.HTTP_204_NO_CONTENT
)
def delete_grade(
    group_id: int,
    student_id: int,
    date: date,
    current_user: models.User = Depends(get_current_active_user),
    db: Session = Depends(get_db),
):
    group = crud.get_group_by_id(db, group_id, current_user.id)
    if not group:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND, detail="Group not found"
        )
    crud.delete_grade(db, group_id, student_id, date)


@app.get("/groups/{group_id}/debtors", response_model=List[schemas.DebtorOut])
def list_debtors(
    group_id: int,
    year: int,
    month: int,
    current_user: models.User = Depends(get_current_active_user),
    db: Session = Depends(get_db),
):
    validate_year_month(year, month)
    group = crud.get_group_by_id(db, group_id, current_user.id)
    if not group:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND, detail="Group not found"
        )
    from calendar import monthrange

    first = date(year, month, 1)
    last = date(year, month, monthrange(year, month)[1])
    students = (
        db.query(models.Student)
        .filter(
            models.Student.group_id == group_id,
            models.Student.status == "active",
        )
        .order_by(models.Student.full_name.asc())
        .all()
    )
    paid_map: dict = {}
    rows = (
        db.query(models.Payment.student_id, models.Payment.amount)
        .filter(
            models.Payment.group_id == group_id,
            models.Payment.date >= first,
            models.Payment.date <= last,
        )
        .all()
    )
    for sid, amount in rows:
        paid_map[sid] = paid_map.get(sid, 0) + amount
    fee = group.monthly_fee or 0
    out = []
    for s in students:
        paid = paid_map.get(s.id, 0)
        due = max(fee - paid, 0)
        out.append(
            schemas.DebtorOut(
                student_id=s.id,
                full_name=s.full_name,
                parent_name=s.parent_name,
                parent_phone=s.parent_phone,
                fee=fee,
                paid=paid,
                due=due,
            )
        )
    return out


@app.get("/groups/{group_id}/schedule", response_model=List[date])
def get_group_schedule(
    group_id: int,
    year: int,
    month: int,
    current_user: models.User = Depends(get_current_active_user),
    db: Session = Depends(get_db),
):
    validate_year_month(year, month)
    group = crud.get_group_by_id(db, group_id, current_user.id)
    if not group:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND, detail="Group not found"
        )
    dates = schedule.get_group_dates_for_month(group.days_of_week, year, month)
    exc_days = {
        e.date for e in crud.get_exception_days_for_month(db, group_id, year, month)
    }
    return [d for d in dates if d not in exc_days]


@app.get(
    "/groups/{group_id}/students/{student_id}/stats",
    response_model=schemas.StudentStats,
)
def student_stats(
    group_id: int,
    student_id: int,
    year: int,
    month: int,
    current_user: models.User = Depends(get_current_active_user),
    db: Session = Depends(get_db),
):
    validate_year_month(year, month)
    group = crud.get_group_by_id(db, group_id, current_user.id)
    if not group:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND, detail="Group not found"
        )
    student = crud.get_student_by_id(db, student_id, group_id)
    if not student:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND, detail="Student not found"
        )
    attendances = crud.get_attendance_for_group_month(db, group_id, year, month)
    total = len([a for a in attendances if a.student_id == student_id])
    present = len(
        [a for a in attendances if a.student_id == student_id and a.status == "present"]
    )
    absent = len(
        [a for a in attendances if a.student_id == student_id and a.status == "absent"]
    )
    excused = len(
        [a for a in attendances if a.student_id == student_id and a.status == "excused"]
    )
    monthly = []
    for m in range(1, 13):
        from calendar import monthrange

        ld = monthrange(year, m)[1]
        ma = [a for a in attendances if a.date.month == m]
        mt = len([a for a in ma if a.student_id == student_id])
        mp = len(
            [a for a in ma if a.student_id == student_id and a.status == "present"]
        )
        monthly.append({"month": m, "total": mt, "present": mp})
    return schemas.StudentStats(
        student_id=student.id,
        full_name=student.full_name,
        total_lessons=total,
        present=present,
        absent=absent,
        excused=excused,
        percentage=round(present / total * 100, 1) if total else 0,
        monthly=monthly,
    )


@app.get("/groups/{group_id}/payments", response_model=List[schemas.PaymentOut])
def list_payments(
    group_id: int,
    year: int,
    month: int,
    current_user: models.User = Depends(get_current_active_user),
    db: Session = Depends(get_db),
):
    validate_year_month(year, month)
    group = crud.get_group_by_id(db, group_id, current_user.id)
    if not group:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND, detail="Group not found"
        )
    from calendar import monthrange

    last_day = monthrange(year, month)[1]
    return (
        db.query(models.Payment)
        .filter(
            models.Payment.group_id == group_id,
            models.Payment.date >= date(year, month, 1),
            models.Payment.date <= date(year, month, last_day),
        )
        .order_by(models.Payment.date.asc())
        .all()
    )


@app.post("/groups/{group_id}/payments", response_model=schemas.PaymentOut)
def create_payment(
    group_id: int,
    payment_in: schemas.PaymentCreate,
    current_user: models.User = Depends(get_current_active_user),
    db: Session = Depends(get_db),
):
    group = crud.get_group_by_id(db, group_id, current_user.id)
    if not group:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND, detail="Group not found"
        )
    student = crud.get_student_by_id(db, payment_in.student_id, group_id)
    if not student:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND, detail="Student not found"
        )
    payment = models.Payment(
        group_id=group_id,
        student_id=payment_in.student_id,
        amount=payment_in.amount,
        date=payment_in.date,
        note=payment_in.note,
    )
    db.add(payment)
    db.commit()
    db.refresh(payment)
    return payment


@app.delete(
    "/groups/{group_id}/payments/{payment_id}", status_code=status.HTTP_204_NO_CONTENT
)
def delete_payment(
    group_id: int,
    payment_id: int,
    current_user: models.User = Depends(get_current_active_user),
    db: Session = Depends(get_db),
):
    group = crud.get_group_by_id(db, group_id, current_user.id)
    if not group:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND, detail="Group not found"
        )
    payment = (
        db.query(models.Payment)
        .filter(models.Payment.id == payment_id, models.Payment.group_id == group_id)
        .first()
    )
    if not payment:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND, detail="Payment not found"
        )
    db.delete(payment)
    db.commit()


@app.get("/superadmin/users", response_model=List[schemas.UserProfileAdmin])
def list_users(
    current_user: models.User = Depends(get_current_superuser),
    db: Session = Depends(get_db),
):
    users = db.query(models.User).order_by(models.User.created_at.desc()).all()
    result = []
    for u in users:
        gc = db.query(models.Group).filter(models.Group.user_id == u.id).count()
        result.append(
            schemas.UserProfileAdmin(
                id=u.id,
                email=u.email,
                username=u.username,
                full_name=u.full_name,
                is_active=u.is_active,
                is_superuser=u.is_superuser,
                telegram_chat_id=u.telegram_chat_id,
                language=u.language,
                created_at=u.created_at,
                group_count=gc,
            )
        )
    return result


@app.get("/superadmin/login-logs", response_model=List[schemas.LoginLogOut])
def login_logs(
    limit: int = 50,
    current_user: models.User = Depends(get_current_superuser),
    db: Session = Depends(get_db),
):
    return (
        db.query(models.LoginLog)
        .order_by(models.LoginLog.id.desc())
        .limit(min(limit, 200))
        .all()
    )


@app.get("/superadmin/login-stats")
def login_stats(
    current_user: models.User = Depends(get_current_superuser),
    db: Session = Depends(get_db),
):
    total = db.query(models.LoginLog).count()
    success = db.query(models.LoginLog).filter(models.LoginLog.success == True).count()
    today = date.today()
    today_count = (
        db.query(models.LoginLog)
        .filter(models.LoginLog.created_at >= today.isoformat())
        .count()
    )
    return {
        "total": total,
        "success": success,
        "failed": total - success,
        "today": today_count,
    }


@app.patch("/superadmin/users/{user_id}", response_model=schemas.UserProfile)
def update_user(
    user_id: int,
    upd: schemas.UserUpdateAdmin,
    current_user: models.User = Depends(get_current_superuser),
    db: Session = Depends(get_db),
):
    user = db.query(models.User).filter(models.User.id == user_id).first()
    if not user:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND, detail="User not found"
        )
    if upd.is_active is not None:
        user.is_active = upd.is_active
    user.is_superuser = user.email.strip().lower() == ADMIN_EMAIL
    db.commit()
    db.refresh(user)
    return user


@app.delete("/superadmin/users/{user_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_user(
    user_id: int,
    current_user: models.User = Depends(get_current_superuser),
    db: Session = Depends(get_db),
):
    if user_id == current_user.id:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST, detail="Cannot delete yourself"
        )
    user = db.query(models.User).filter(models.User.id == user_id).first()
    if not user:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND, detail="User not found"
        )
    db.delete(user)
    db.commit()


UPLOAD_DIR = os.path.join(
    os.path.dirname(os.path.abspath(__file__)), "..", "uploads", "receipts"
)


@app.get("/groups/{group_id}/receipts", response_model=List[schemas.ReceiptOut])
def list_receipts(
    group_id: int,
    current_user: models.User = Depends(get_current_active_user),
    db: Session = Depends(get_db),
):
    receipts = (
        db.query(models.PaymentReceipt)
        .filter(models.PaymentReceipt.group_id == group_id)
        .order_by(models.PaymentReceipt.created_at.desc())
        .limit(100)
        .all()
    )
    return receipts


@app.post(
    "/groups/{group_id}/receipts/{receipt_id}/confirm",
    response_model=schemas.PaymentOut,
)
def confirm_receipt(
    group_id: int,
    receipt_id: int,
    payload: schemas.ReceiptConfirm,
    current_user: models.User = Depends(get_current_active_user),
    db: Session = Depends(get_db),
):
    receipt = (
        db.query(models.PaymentReceipt)
        .filter(
            models.PaymentReceipt.id == receipt_id,
            models.PaymentReceipt.group_id == group_id,
        )
        .first()
    )
    if not receipt:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND, detail="Receipt not found"
        )
    if receipt.status != "pending":
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST, detail="Receipt is not pending"
        )
    from datetime import datetime

    receipt.status = "confirmed"
    receipt.amount = payload.amount
    receipt.confirmed_at = datetime.now()
    receipt.confirmed_by = current_user.id
    payment = models.Payment(
        group_id=group_id,
        student_id=receipt.student_id,
        amount=payload.amount,
        date=payload.date,
        note=f"Chek #{receipt.id}",
    )
    db.add(payment)
    db.commit()
    db.refresh(payment)
    if BOT_ENABLED:
        try:
            telegram_bot.notify_receipt_status(receipt.id, "confirmed")
        except Exception:
            pass
    return payment


@app.post(
    "/groups/{group_id}/receipts/{receipt_id}/reject",
    status_code=status.HTTP_200_OK,
)
def reject_receipt(
    group_id: int,
    receipt_id: int,
    current_user: models.User = Depends(get_current_active_user),
    db: Session = Depends(get_db),
):
    receipt = (
        db.query(models.PaymentReceipt)
        .filter(
            models.PaymentReceipt.id == receipt_id,
            models.PaymentReceipt.group_id == group_id,
        )
        .first()
    )
    if not receipt:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND, detail="Receipt not found"
        )
    if receipt.status != "pending":
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST, detail="Receipt is not pending"
        )
    from datetime import datetime

    receipt.status = "rejected"
    receipt.confirmed_at = datetime.now()
    receipt.confirmed_by = current_user.id
    db.commit()
    if BOT_ENABLED:
        try:
            telegram_bot.notify_receipt_status(receipt.id, "rejected")
        except Exception:
            pass
    return {"ok": True}


@app.get("/receipts/{receipt_id}/photo")
def receipt_photo(
    receipt_id: int,
    current_user: models.User = Depends(get_current_active_user),
    db: Session = Depends(get_db),
):
    from fastapi.responses import FileResponse

    receipt = (
        db.query(models.PaymentReceipt)
        .filter(models.PaymentReceipt.id == receipt_id)
        .first()
    )
    if not receipt or not receipt.photo_path or not os.path.exists(receipt.photo_path):
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND, detail="Photo not found"
        )
    return FileResponse(receipt.photo_path)
