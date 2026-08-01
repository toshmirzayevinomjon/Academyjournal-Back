import asyncio
import io
import logging
import os
import time
from datetime import date, datetime, timedelta
from typing import Optional

from sqlalchemy.orm import Session as DBSession

from telegram import (
    Update,
    BotCommand,
    InlineKeyboardButton,
    InlineKeyboardMarkup,
    KeyboardButton,
    ReplyKeyboardMarkup,
)
from telegram.ext import (
    Application,
    CallbackQueryHandler,
    CommandHandler,
    ContextTypes,
    MessageHandler,
    filters,
)
from telegram.constants import ParseMode

from . import models
from .database import SessionLocal, engine
from .schedule import get_group_dates_for_month

logger = logging.getLogger(__name__)

TOKEN = os.environ.get("TELEGRAM_BOT_TOKEN", "")
MONTHLY_FEE_UZS = int(os.environ.get("MONTHLY_FEE_UZS", "500000"))
WEBHOOK_URL = os.environ.get("TELEGRAM_WEBHOOK_URL", "")
WEBHOOK_PORT = int(os.environ.get("TELEGRAM_WEBHOOK_PORT", "8443"))
WEBHOOK_SECRET = os.environ.get("TELEGRAM_WEBHOOK_SECRET", "kundalik-secret")

MONTH_NAMES_UZ = [
    "Yanvar",
    "Fevral",
    "Mart",
    "Aprel",
    "May",
    "Iyun",
    "Iyul",
    "Avgust",
    "Sentabr",
    "Oktyabr",
    "Noyabr",
    "Dekabr",
]

_bot_app: Optional[Application] = None
_link_codes: dict = {}
_pending_parents: dict = {}  # chat_id -> {"phone": str, "students": list}
_pay_selection: dict = {}  # chat_id -> student_id (oxirgi /pay tanlovi)

COMMANDS_MENU = [
    ("start", "Botni boshlash"),
    ("help", "Yordam"),
    ("groups", "Guruhlarim"),
    ("today", "Bugungi davomat"),
    ("stats", "Oylik statistika"),
    ("student", "O'quvchi qidirish"),
    ("report", "Hisobot (Excel)"),
    ("reportpdf", "Hisobot (PDF)"),
    ("mark", "Davomat belgilash"),
    ("topic", "Mavzu qo'shish"),
    ("rating", "Guruh reytingi"),
    ("payments", "To'lov hisoboti"),
    ("paymentlist", "To'lamaganlar"),
    ("paymenthistory", "To'lov tarixi"),
    ("phone", "Ota-ona ulanish"),
    ("mystudent", "Farzandim haqida"),
    ("pay", "To'lov qilish (chek)"),
    ("chart", "O'quvchi grafigi"),
    ("addgroup", "Guruh qo'shish"),
    ("addstudent", "O'quvchi qo'shish"),
    ("editstudent", "O'quvchini tahrirlash"),
    ("removestudent", "O'quvchini o'chirish"),
    ("archive", "Guruhni arxivlash"),
    ("setfee", "Oylik to'lov narxi"),
    ("setcard", "Karta raqamini o'rnatish"),
    ("broadcast", "Barchaga e'lon"),
]


def teacher_keyboard() -> ReplyKeyboardMarkup:
    return ReplyKeyboardMarkup(
        [
            [KeyboardButton("📋 Bugungi davomat"), KeyboardButton("📊 Statistika")],
            [KeyboardButton("📚 Guruhlarim"), KeyboardButton("💰 To'lov hisoboti")],
            [KeyboardButton("🧾 To'lov qilish"), KeyboardButton("⭐ Reyting")],
            [KeyboardButton("📄 Hisobot (Excel)"), KeyboardButton("❓ Yordam")],
        ],
        resize_keyboard=True,
        input_field_placeholder="Komandani tanlang...",
    )


def parent_keyboard() -> ReplyKeyboardMarkup:
    return ReplyKeyboardMarkup(
        [
            [KeyboardButton("👶 Farzandim"), KeyboardButton("💳 To'lov qilish")],
            [KeyboardButton("💳 To'lov tarixi"), KeyboardButton("❓ Yordam")],
        ],
        resize_keyboard=True,
        input_field_placeholder="Komandani tanlang...",
    )


def normalize_phone(p: str) -> str:
    return "".join(c for c in p if c.isdigit())


def generate_link_code(user_id: int) -> str:
    import random, string

    code = "".join(random.choices(string.ascii_uppercase + string.digits, k=6))
    _link_codes[code] = {
        "user_id": user_id,
        "expires": datetime.now() + timedelta(minutes=5),
    }
    return code


def get_link_code_user(code: str) -> Optional[int]:
    entry = _link_codes.get(code)
    if not entry:
        return None
    if datetime.now() > entry["expires"]:
        _link_codes.pop(code, None)
        return None
    return entry["user_id"]


def get_user_by_chat_id(db: DBSession, chat_id: int) -> Optional[models.User]:
    return (
        db.query(models.User)
        .filter(
            models.User.telegram_chat_id == str(chat_id),
            models.User.is_active == True,
        )
        .first()
    )


def get_or_create_temp_user(db: DBSession, chat_id: int) -> Optional[models.User]:
    return (
        db.query(models.User)
        .filter(
            models.User.telegram_chat_id == str(chat_id),
        )
        .first()
    )


def format_date(d: date) -> str:
    return f"{d.day}/{d.month}"


def format_weekday(d: date) -> str:
    days = ["Du", "Se", "Chor", "Pay", "Ju", "Shan", "Yak"]
    return days[d.weekday()]


async def cmd_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    chat_id = update.effective_chat.id
    db: DBSession = SessionLocal()
    try:
        user = get_user_by_chat_id(db, chat_id)
        if user:
            await update.message.reply_text(
                f"Assalomu alaykum, {user.full_name}! ✅\n\n"
                f"Sizning hisobingiz Telegram ga ulangan.\n"
                f"Quyidagi menyudan komanda tanlang yoki /help ni bosing:\n\n"
                f"/groups - Guruhlarim\n"
                f"/today - Bugungi davomat\n"
                f"/stats - Oylik statistika\n"
                f"/student Ismi - O'quvchi qidirish\n"
                f"/report - Hisobot\n"
                f"/payments - To'lov hisoboti",
                reply_markup=teacher_keyboard(),
            )
            return
        parent_links = (
            db.query(models.ParentLink)
            .filter(models.ParentLink.chat_id == str(chat_id))
            .all()
        )
        if parent_links:
            await update.message.reply_text(
                "Assalomu alaykum! 👋\n"
                "Siz ota-ona sifatida ulangansiz.\n"
                "Menyudan komanda tanlang:\n\n"
                "/mystudent - Farzandingiz haqida\n"
                "/pay - To'lov qilish\n"
                "/paymenthistory - To'lov tarixi",
                reply_markup=parent_keyboard(),
            )
            return
        link_code = context.args[0].strip().upper() if context.args else None
        if link_code:
            uid = get_link_code_user(link_code)
            if uid:
                u = db.query(models.User).filter(models.User.id == uid).first()
                if u:
                    u.telegram_chat_id = str(chat_id)
                    db.commit()
                    _link_codes.pop(link_code, None)
                    await update.message.reply_text(
                        f"✅ <b>Hisobingiz muvaffaqiyatli ulandi!</b>\n\n"
                        f"Xush kelibsiz, {u.full_name}!\n"
                        f"Endi quyidagi buyruqlardan foydalaning:\n\n"
                        f"/groups - Guruhlarim\n"
                        f"/today - Bugungi davomat\n"
                        f"/stats - Oylik statistika\n"
                        f"/student Ismi - O'quvchi qidirish\n"
                        f"/report - Hisobot\n"
                        f"/payments - To'lov hisoboti",
                        parse_mode=ParseMode.HTML,
                    )
                    return
            await update.message.reply_text("❌ Kod noto'g'ri yoki muddati o'tgan.")
            return
        await update.message.reply_text(
            "Assalomu alaykum! 👋\n\n"
            "Bu Kundalik LMS davomat tizimining Telegram boti.\n\n"
            "<b>O'qituvchilar</b> uchun:\n"
            '1. Web sayt → Sozlamalar → <b>"Telegramga ulanish"</b> tugmasini bosing\n'
            "2. Telegram avtomatik ulanadi\n\n"
            "<b>Ota-onalar</b> uchun:\n"
            "Farzandingizning telefon raqamini yuboring:\n"
            "/phone 998901234567\n\n"
            "<i>Yordam: /help</i>",
            parse_mode=ParseMode.HTML,
        )
    finally:
        db.close()


async def cmd_link(update: Update, context: ContextTypes.DEFAULT_TYPE):
    chat_id = update.effective_chat.id
    if not context.args:
        await update.message.reply_text("Kodni kiriting. Masalan: /link ABC123")
        return
    code = context.args[0].strip().upper()
    user_id = get_link_code_user(code)
    if not user_id:
        await update.message.reply_text(
            "❌ Kod noto'g'ri yoki muddati o'tgan. Web saytdan yangi kod oling."
        )
        return
    _link_codes.pop(code, None)
    db: DBSession = SessionLocal()
    try:
        user = db.query(models.User).filter(models.User.id == user_id).first()
        if not user:
            await update.message.reply_text("❌ Foydalanuvchi topilmadi.")
            return
        user.telegram_chat_id = str(chat_id)
        db.commit()
        await update.message.reply_text(
            f"✅ Hisobingiz muvaffaqiyatli ulandi, {user.full_name}!\n\n"
            f"Endi quyidagi buyruqlardan foydalanishingiz mumkin:\n"
            f"/groups - Guruhlarim\n"
            f"/today - Bugungi davomat\n"
            f"/stats - Oylik statistika\n"
            f"/student Ismi - O'quvchi qidirish\n"
            f"/report - Hisobot\n"
            f"/payments - To'lov hisoboti"
        )
    finally:
        db.close()


async def cmd_help(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text(
        "/start - Botni ulash\n"
        "/groups - Guruhlar ro'yxati\n"
        "/today - Bugungi davomat\n"
        "/stats - Oylik statistika\n"
        "/student Ismi - O'quvchi qidirish\n"
        "/report - Hisobot (Excel)\n"
        "/reportpdf - Hisobot (PDF, namuna: /reportpdf 2026-07)\n"
        "/payments - To'lov hisoboti\n"
        "/phone 998901234567 - Ota-onalar uchun ulanish\n"
        "/mystudent - Farzandingiz haqida ma'lumot\n"
        "/pay - To'lov qilish (karta raqami, chek yuborish)\n"
        "/paymenthistory - To'lov tarixi\n"
        "/paymentlist - To'lamaganlar ro'yxati\n"
        "/rating - Guruh reytingi\n"
        "/chart o'quvchi_raqami - O'quvchi grafigi\n"
        "/topic 1 Mavzu --homework 'vazifa' - Mavzu qo'shish\n"
        "/broadcast Xabar - Barchaga e'lon (admin)\n"
        "/editstudent 1 5 --holat inactive - O'quvchini tahrirlash\n"
        "/removestudent 1 5 - O'quvchini o'chirish\n"
        "/archive 1 - Guruhni arxivlash/qaytarish\n"
        "/setfee 600000 - Oylik to'lov narxini belgilash\n"
        "/topstudents - Oyning eng yaxshi o'quvchilari\n"
        "/schedule - Haftalik dars jadvali\n"
        "/topic - Bugungi mavzu\n"
        "/mark - Davomat belgilash\n"
        "/addstudent - O'quvchi qo'shish\n"
        "/addgroup - Guruh qo'shish\n"
        "/chart - Statistika grafigi\n"
        "/lang - Tilni o'zgartirish\n"
        "/admin - Admin boshqaruvi\n"
        "/help - Yordam"
    )


async def cmd_phone(update: Update, context: ContextTypes.DEFAULT_TYPE):
    chat_id = update.effective_chat.id
    if not context.args:
        await update.message.reply_text(
            "Telefon raqamingizni kiriting. Masalan: /phone 998901234567"
        )
        return
    phone = normalize_phone(context.args[0])
    if len(phone) < 10:
        await update.message.reply_text(
            "❌ Noto'g'ri raqam. Masalan: /phone 998901234567"
        )
        return
    db: DBSession = SessionLocal()
    try:
        students = (
            db.query(models.Student)
            .filter(
                models.Student.parent_phone.isnot(None),
                models.Student.parent_phone != "",
            )
            .all()
        )
        matches = [
            s
            for s in students
            if phone in normalize_phone(s.parent_phone)
            or normalize_phone(s.parent_phone).endswith(phone)
            or phone.endswith(normalize_phone(s.parent_phone))
        ]
        if not matches:
            existing = (
                db.query(models.ParentLink)
                .filter(models.ParentLink.chat_id == str(chat_id))
                .first()
            )
            if existing:
                await update.message.reply_text(
                    f"Siz allaqachon {existing.student.full_name} ga ulangansiz.\n"
                    f"/mystudent - ma'lumot\n"
                    f"/phone 998901234567 - boshqa farzandga ulanish"
                )
                return
            await update.message.reply_text(
                "❌ Bu raqam bilan bog'langan o'quvchi topilmadi.\n"
                "Iltimos, farzandingiz o'qituvchisidan to'g'ri telefon raqamini oling."
            )
            return
        if len(matches) == 1:
            s = matches[0]
            existing = (
                db.query(models.ParentLink)
                .filter(models.ParentLink.student_id == s.id)
                .first()
            )
            if existing:
                db.query(models.ParentLink).filter(
                    models.ParentLink.chat_id == str(chat_id)
                ).delete()
                db.commit()
            pl = models.ParentLink(student_id=s.id, chat_id=str(chat_id), phone=phone)
            db.add(pl)
            db.commit()
            await update.message.reply_text(
                f"✅ <b>Muvaffaqiyatli ulandi!</b>\n\n"
                f"Farzandingiz: {s.full_name}\n"
                f"Endi farzandingiz darsga kelmaganda xabar olasiz.",
                parse_mode=ParseMode.HTML,
                reply_markup=parent_keyboard(),
            )
        else:
            _pending_parents[chat_id] = {"phone": phone, "students": matches}
            lines = ["<b>Bir nechta o'quvchi topildi.</b>\n"]
            for i, s in enumerate(matches, 1):
                g = db.query(models.Group).filter(models.Group.id == s.group_id).first()
                lines.append(f"{i}. {s.full_name} — {g.name if g else '?'}")
            lines.append(f"\nTanlang: /select 1 dan {len(matches)} gacha")
            await update.message.reply_text("\n".join(lines), parse_mode=ParseMode.HTML)
    finally:
        db.close()


async def cmd_select(update: Update, context: ContextTypes.DEFAULT_TYPE):
    chat_id = update.effective_chat.id
    if chat_id not in _pending_parents:
        await update.message.reply_text("Avval /phone orqali raqamingizni yuboring.")
        return
    if not context.args:
        await update.message.reply_text("Raqam kiriting. Masalan: /select 1")
        return
    try:
        idx = int(context.args[0]) - 1
        data = _pending_parents[chat_id]
        if idx < 0 or idx >= len(data["students"]):
            await update.message.reply_text(
                f"❌ 1 dan {len(data['students'])} gacha tanlang."
            )
            return
        s = data["students"][idx]
        db: DBSession = SessionLocal()
        try:
            existing = (
                db.query(models.ParentLink)
                .filter(models.ParentLink.student_id == s.id)
                .first()
            )
            if existing:
                db.query(models.ParentLink).filter(
                    models.ParentLink.chat_id == str(chat_id)
                ).delete()
                db.commit()
            pl = models.ParentLink(
                student_id=s.id, chat_id=str(chat_id), phone=data["phone"]
            )
            db.add(pl)
            db.commit()
            _pending_parents.pop(chat_id, None)
            await update.message.reply_text(
                f"✅ <b>Muvaffaqiyatli ulandi!</b>\n\n"
                f"Farzandingiz: {s.full_name}\n"
                f"Endi farzandingiz darsga kelmaganda xabar olasiz.",
                parse_mode=ParseMode.HTML,
            )
        finally:
            db.close()
    except ValueError:
        await update.message.reply_text("❌ Noto'g'ri raqam.")


async def cmd_mystudent(update: Update, context: ContextTypes.DEFAULT_TYPE):
    chat_id = update.effective_chat.id
    db: DBSession = SessionLocal()
    try:
        link = (
            db.query(models.ParentLink)
            .filter(models.ParentLink.chat_id == str(chat_id))
            .first()
        )
        if not link:
            await update.message.reply_text(
                "Siz hali farzandingizga ulanmagansiz.\n"
                "/phone 998901234567 orqali ulaning."
            )
            return
        s = (
            db.query(models.Student)
            .filter(models.Student.id == link.student_id)
            .first()
        )
        if not s:
            await update.message.reply_text("❌ O'quvchi topilmadi.")
            return
        g = db.query(models.Group).filter(models.Group.id == s.group_id).first()
        today = date.today()
        from .schedule import get_group_dates_for_month

        dates = (
            get_group_dates_for_month(
                g.days_of_week if g else [], today.year, today.month
            )
            if g
            else []
        )
        present = absent = excused = 0
        for d in dates:
            att = (
                db.query(models.Attendance)
                .filter(
                    models.Attendance.student_id == s.id,
                    models.Attendance.date == d,
                )
                .first()
            )
            if att:
                if att.status == "present":
                    present += 1
                elif att.status == "absent":
                    absent += 1
                elif att.status == "excused":
                    excused += 1
        pct = (
            round(present / (present + absent + excused) * 100)
            if (present + absent + excused)
            else 0
        )
        await update.message.reply_text(
            f"👤 <b>{s.full_name}</b>\n"
            f"   📚 {g.name if g else '—'}\n"
            f"   📞 {s.phone or '—'}\n\n"
            f"📊 {MONTH_NAMES_UZ[today.month - 1]}:\n"
            f"   ✅ Kelgan: {present}\n"
            f"   ❌ Yo'q: {absent}\n"
            f"   ⏳ Sababli: {excused}\n"
            f"   🎯 Davomat: {pct}%",
            parse_mode=ParseMode.HTML,
        )
    finally:
        db.close()


async def cmd_pay(update: Update, context: ContextTypes.DEFAULT_TYPE):
    chat_id = update.effective_chat.id
    db: DBSession = SessionLocal()
    try:
        links = (
            db.query(models.ParentLink)
            .filter(models.ParentLink.chat_id == str(chat_id))
            .all()
        )
        if not links:
            await update.message.reply_text(
                "To'lov uchun avval farzandingizga ulaning:\n/phone 998901234567"
            )
            return
        students = [
            db.query(models.Student).filter(models.Student.id == l.student_id).first()
            for l in links
        ]
        students = [s for s in students if s]
        if not students:
            await update.message.reply_text("❌ O'quvchi topilmadi.")
            return
        idx = 0
        if context.args:
            try:
                idx = int(context.args[0]) - 1
            except ValueError:
                await update.message.reply_text("❌ Noto'g'ri raqam. Masalan: /pay 2")
                return
            if idx < 0 or idx >= len(students):
                await update.message.reply_text(
                    f"❌ 1 dan {len(students)} gacha tanlang."
                )
                return
        if len(students) > 1 and not context.args:
            lines = [
                "O'quvchingizni tanlang:",
            ]
            for i, s in enumerate(students, 1):
                g = db.query(models.Group).filter(models.Group.id == s.group_id).first()
                lines.append(f"{i}. {s.full_name} ({g.name if g else '—'})")
            lines.append(f"\nMasalan: /pay 1")
            await update.message.reply_text("\n".join(lines))
            return
        s = students[idx]
        _pay_selection[str(chat_id)] = s.id
        g = db.query(models.Group).filter(models.Group.id == s.group_id).first()
        teacher = (
            db.query(models.User).filter(models.User.id == g.user_id).first()
            if g
            else None
        )
        card = teacher.card_number if teacher else None
        fee = (
            teacher.monthly_fee if teacher and teacher.monthly_fee else MONTHLY_FEE_UZS
        )
        if not card:
            await update.message.reply_text(
                "❌ To'lov kartasi hozircha o'rnatilmagan. O'qituvchi bilan bog'laning."
            )
            return
        await update.message.reply_text(
            f"💳 <b>To'lov qilish</b>\n\n"
            f"👤 {s.full_name}\n"
            f"📚 {g.name if g else '—'}\n"
            f"💵 {fee:,} so'm\n\n"
            f"<b>Karta raqami:</b> <code>{card}</code>\n\n"
            f"1️⃣ Pulni karta raqamiga o'tkazing\n"
            f"2️⃣ Chekni (o'tkazma kvitansiyasini) rasm sifatida yuboring\n"
            f"3️⃣ O'qituvchi tasdiqlagach, sizga xabar beramiz",
            parse_mode=ParseMode.HTML,
            reply_markup=InlineKeyboardMarkup(
                [
                    [
                        InlineKeyboardButton(
                            "✅ Chek yubordim",
                            callback_data="receipt:ready",
                        )
                    ]
                ]
            ),
        )
    finally:
        db.close()


async def cmd_receipt_photo(update: Update, context: ContextTypes.DEFAULT_TYPE):
    chat_id = update.effective_chat.id
    db: DBSession = SessionLocal()
    try:
        links = (
            db.query(models.ParentLink)
            .filter(models.ParentLink.chat_id == str(chat_id))
            .all()
        )
        if not links:
            await update.message.reply_text(
                "Chek yuborish uchun avval farzandingizga ulaning:\n/phone 998901234567"
            )
            return
        selected = _pay_selection.get(str(chat_id))
        student = None
        if selected:
            student = (
                db.query(models.Student).filter(models.Student.id == selected).first()
            )
            if not student or not any(l.student_id == student.id for l in links):
                student = None
        if not student:
            for l in links:
                student = (
                    db.query(models.Student)
                    .filter(models.Student.id == l.student_id)
                    .first()
                )
                if student:
                    break
        if not student:
            await update.message.reply_text("❌ O'quvchi topilmadi.")
            return
        photo = update.message.photo[-1]
        file = await context.bot.get_file(photo.file_id)
        upload_dir = os.path.join(
            os.path.dirname(os.path.abspath(__file__)), "..", "uploads", "receipts"
        )
        os.makedirs(upload_dir, exist_ok=True)
        tmp_path = os.path.join(
            upload_dir, f"tmp_{int(time.time() * 1000)}_{photo.file_unique_id}.jpg"
        )
        await file.download_to_drive(tmp_path)
        receipt = models.PaymentReceipt(
            group_id=student.group_id,
            student_id=student.id,
            chat_id=str(chat_id),
            photo_path=tmp_path,
            status="pending",
            note=update.message.caption,
        )
        db.add(receipt)
        db.commit()
        db.refresh(receipt)
        final_path = os.path.join(upload_dir, f"receipt_{receipt.id}.jpg")
        try:
            os.rename(tmp_path, final_path)
            receipt.photo_path = final_path
            db.commit()
        except Exception:
            final_path = tmp_path
        await update.message.reply_text(
            f"✅ <b>Chek qabul qilindi!</b>\n\n"
            f"👤 {student.full_name}\n"
            f"📅 {datetime.now().strftime('%d.%m.%Y %H:%M')}\n\n"
            f"O'qituvchi tasdiqlagach, sizga xabar beramiz. Rahmat! 🙂",
            parse_mode=ParseMode.HTML,
        )
        group = (
            db.query(models.Group).filter(models.Group.id == student.group_id).first()
        )
        teacher = (
            db.query(models.User).filter(models.User.id == group.user_id).first()
            if group
            else None
        )
        if teacher and teacher.telegram_chat_id:
            try:
                with open(final_path, "rb") as f:
                    await context.bot.send_photo(
                        chat_id=int(teacher.telegram_chat_id),
                        photo=f,
                        caption=(
                            f"🧾 <b>Yangi to'lov cheki keldi</b>\n\n"
                            f"👤 {student.full_name}\n"
                            f"📚 {group.name if group else '—'}\n"
                            f"🆔 Chek #{receipt.id}\n\n"
                            f"Tasdiqlash: Web sayt → To'lov hisoboti → Kutilayotgan cheklar"
                        ),
                        parse_mode=ParseMode.HTML,
                    )
            except Exception as e:
                logger.error(f"Failed to notify teacher of receipt: {e}")
    finally:
        db.close()


async def cmd_groups(update: Update, context: ContextTypes.DEFAULT_TYPE):
    chat_id = update.effective_chat.id
    db: DBSession = SessionLocal()
    try:
        user = get_user_by_chat_id(db, chat_id)
        if not user:
            await update.message.reply_text("Avval /start orqali hisobingizni ulang.")
            return
        groups = (
            db.query(models.Group)
            .filter(
                models.Group.user_id == user.id,
                models.Group.is_archived == False,
            )
            .all()
        )
        if not groups:
            await update.message.reply_text("Guruhlar mavjud emas.")
            return
        lines = ["<b>Mening guruhlarim:</b>\n"]
        for g in groups:
            sc = (
                db.query(models.Student)
                .filter(
                    models.Student.group_id == g.id,
                    models.Student.status == "active",
                )
                .count()
            )
            days_str = ", ".join(g.days_of_week or [])
            lines.append(
                f"📚 <b>{g.name}</b>\n"
                f"   👥 {sc} ta o'quvchi | {days_str}\n"
                f"   🕐 {g.lesson_time.strftime('%H:%M') if g.lesson_time else '--:--'}"
            )
        await update.message.reply_text("\n".join(lines), parse_mode=ParseMode.HTML)
    finally:
        db.close()


async def cmd_today(update: Update, context: ContextTypes.DEFAULT_TYPE):
    chat_id = update.effective_chat.id
    db: DBSession = SessionLocal()
    try:
        user = get_user_by_chat_id(db, chat_id)
        if not user:
            await update.message.reply_text("Avval /start orqali hisobingizni ulang.")
            return
        today = date.today()
        groups = (
            db.query(models.Group)
            .filter(
                models.Group.user_id == user.id,
                models.Group.is_archived == False,
            )
            .all()
        )

        lines = [f"<b>Bugun: {today.day}/{today.month}/{today.year}</b>\n"]
        has_data = False
        for g in groups:
            weekday_name = ["MON", "TUE", "WED", "THU", "FRI", "SAT", "SUN"][
                today.weekday()
            ]
            if weekday_name not in (g.days_of_week or []):
                continue
            exc = (
                db.query(models.ExceptionDay)
                .filter(
                    models.ExceptionDay.group_id == g.id,
                    models.ExceptionDay.date == today,
                )
                .first()
            )
            if exc:
                lines.append(f"📚 {g.name} - <b>ISTI SN</b> ({exc.reason or ''})")
                continue
            students = (
                db.query(models.Student)
                .filter(
                    models.Student.group_id == g.id,
                    models.Student.status == "active",
                )
                .all()
            )
            if not students:
                continue
            att_map = {}
            for s in students:
                att = (
                    db.query(models.Attendance)
                    .filter(
                        models.Attendance.group_id == g.id,
                        models.Attendance.student_id == s.id,
                        models.Attendance.date == today,
                    )
                    .first()
                )
                att_map[s.id] = att.status if att else None
            present = sum(1 for v in att_map.values() if v == "present")
            absent = sum(1 for v in att_map.values() if v == "absent")
            excused = sum(1 for v in att_map.values() if v == "excused")
            not_set = sum(1 for v in att_map.values() if v is None)
            total = len(students)
            if not_set < total:
                has_data = True
                lines.append(
                    f"\n📚 <b>{g.name}</b>\n"
                    f"   ✅ Kelgan: {present}\n"
                    f"   ❌ Yo'q: {absent}\n"
                    f"   ⏳ Sababli: {excused}\n"
                    f"   📊 {total} ta o'quvchidan {present}/{total}"
                )

        if not has_data:
            lines.append("\nBugun dars kuni emas yoki davomat hali belgilanmagan.")
        await update.message.reply_text("\n".join(lines), parse_mode=ParseMode.HTML)
    finally:
        db.close()


async def cmd_stats(update: Update, context: ContextTypes.DEFAULT_TYPE):
    chat_id = update.effective_chat.id
    db: DBSession = SessionLocal()
    try:
        user = get_user_by_chat_id(db, chat_id)
        if not user:
            await update.message.reply_text("Avval /start orqali hisobingizni ulang.")
            return
        today = date.today()
        groups = (
            db.query(models.Group)
            .filter(
                models.Group.user_id == user.id,
                models.Group.is_archived == False,
            )
            .all()
        )

        for g in groups:
            dates = get_group_dates_for_month(g.days_of_week, today.year, today.month)
            if not dates:
                continue
            students = (
                db.query(models.Student)
                .filter(
                    models.Student.group_id == g.id,
                    models.Student.status == "active",
                )
                .all()
            )
            if not students:
                continue
            total_lessons = 0
            total_present = 0
            total_absent = 0
            total_excused = 0
            for d in dates:
                exc = (
                    db.query(models.ExceptionDay)
                    .filter(
                        models.ExceptionDay.group_id == g.id,
                        models.ExceptionDay.date == d,
                    )
                    .first()
                )
                if exc:
                    continue
                total_lessons += 1
                for s in students:
                    att = (
                        db.query(models.Attendance)
                        .filter(
                            models.Attendance.group_id == g.id,
                            models.Attendance.student_id == s.id,
                            models.Attendance.date == d,
                        )
                        .first()
                    )
                    if att:
                        if att.status == "present":
                            total_present += 1
                        elif att.status == "absent":
                            total_absent += 1
                        elif att.status == "excused":
                            total_excused += 1
            if total_lessons > 0:
                pct = (
                    round(total_present / (total_lessons * len(students)) * 100)
                    if students
                    else 0
                )
                await update.message.reply_text(
                    f"📚 <b>{g.name}</b>\n"
                    f"   📅 {MONTH_NAMES_UZ[today.month - 1]} {today.year}\n"
                    f"   📊 {total_lessons} ta dars kuni\n"
                    f"   ✅ Kelgan: {total_present}\n"
                    f"   ❌ Yo'q: {total_absent}\n"
                    f"   ⏳ Sababli: {total_excused}\n"
                    f"   🎯 Davomat: {pct}%",
                    parse_mode=ParseMode.HTML,
                )
        await update.message.reply_text("Barcha guruhlar statistikasi yuqorida.")
    finally:
        db.close()


async def cmd_student(update: Update, context: ContextTypes.DEFAULT_TYPE):
    chat_id = update.effective_chat.id
    db: DBSession = SessionLocal()
    try:
        user = get_user_by_chat_id(db, chat_id)
        if not user:
            await update.message.reply_text("Avval /start orqali hisobingizni ulang.")
            return
        query = " ".join(context.args) if context.args else ""
        if not query:
            await update.message.reply_text("Ism kiriting. Masalan: /student Alisher")
            return
        groups = db.query(models.Group).filter(models.Group.user_id == user.id).all()
        group_ids = [g.id for g in groups]
        students = (
            db.query(models.Student)
            .filter(
                models.Student.group_id.in_(group_ids),
                models.Student.full_name.ilike(f"%{query}%"),
            )
            .limit(5)
            .all()
        )
        if not students:
            await update.message.reply_text("O'quvchi topilmadi.")
            return
        today = date.today()
        for st in students:
            group = (
                db.query(models.Group).filter(models.Group.id == st.group_id).first()
            )
            dates = get_group_dates_for_month(
                group.days_of_week if group else [], today.year, today.month
            )
            present = absent = excused = 0
            for d in dates:
                exc = (
                    db.query(models.ExceptionDay)
                    .filter(
                        models.ExceptionDay.group_id == st.group_id,
                        models.ExceptionDay.date == d,
                    )
                    .first()
                )
                if exc:
                    continue
                att = (
                    db.query(models.Attendance)
                    .filter(
                        models.Attendance.student_id == st.id,
                        models.Attendance.date == d,
                    )
                    .first()
                )
                if att:
                    if att.status == "present":
                        present += 1
                    elif att.status == "absent":
                        absent += 1
                    elif att.status == "excused":
                        excused += 1
            total = present + absent + excused
            pct = round(present / total * 100) if total else 0
            await update.message.reply_text(
                f"👤 <b>{st.full_name}</b>\n"
                f"   📚 {group.name if group else '—'}\n"
                f"   📞 {st.phone or '—'}\n"
                f"   👪 {st.parent_name or '—'} ({st.parent_phone or '—'})\n"
                f"   📊 {MONTH_NAMES_UZ[today.month - 1]}:\n"
                f"   ✅ {present} | ❌ {absent} | ⏳ {excused}\n"
                f"   🎯 Davomat: {pct}%",
                parse_mode=ParseMode.HTML,
            )
    finally:
        db.close()


async def cmd_report(update: Update, context: ContextTypes.DEFAULT_TYPE):
    chat_id = update.effective_chat.id
    db: DBSession = SessionLocal()
    try:
        user = get_user_by_chat_id(db, chat_id)
        if not user:
            await update.message.reply_text("Avval /start orqali hisobingizni ulang.")
            return
        await update.message.reply_text("Hisobot tayyorlanmoqda, biroz kuting...")
        from calendar import monthrange
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        import io

        today = date.today()
        groups = (
            db.query(models.Group)
            .filter(
                models.Group.user_id == user.id,
                models.Group.is_archived == False,
            )
            .all()
        )
        for g in groups:
            try:
                dates = get_group_dates_for_month(
                    g.days_of_week, today.year, today.month
                )
                students = (
                    db.query(models.Student)
                    .filter(models.Student.group_id == g.id)
                    .order_by(models.Student.id)
                    .all()
                )
                attendances = (
                    db.query(models.Attendance)
                    .filter(
                        models.Attendance.group_id == g.id,
                        models.Attendance.date >= date(today.year, today.month, 1),
                        models.Attendance.date
                        <= date(
                            today.year,
                            today.month,
                            monthrange(today.year, today.month)[1],
                        ),
                    )
                    .all()
                )
                att_map = {}
                for a in attendances:
                    att_map.setdefault(a.student_id, {})[a.date.isoformat()] = a.status
                wb = Workbook()
                ws = wb.active
                hf = Font(bold=True, color="FFFFFF", size=11)
                hfill = PatternFill(
                    start_color="059669", end_color="0D9488", fill_type="solid"
                )
                thin = Border(
                    left=Side(style="thin", color="D1D5DB"),
                    right=Side(style="thin", color="D1D5DB"),
                    top=Side(style="thin", color="D1D5DB"),
                    bottom=Side(style="thin", color="D1D5DB"),
                )
                ws.cell(1, 1, "O'quvchi").font = hf
                ws.cell(1, 1).fill = hfill
                ws.cell(1, 1).border = thin
                for i, d in enumerate(dates):
                    cell = ws.cell(1, i + 2, d.strftime("%d.%m"))
                    cell.font = hf
                    cell.fill = hfill
                    cell.border = thin
                    cell.alignment = Alignment(horizontal="center")
                col = len(dates) + 2
                for h in ["Kelgan", "Foiz"]:
                    c = ws.cell(1, col, h)
                    c.font = hf
                    c.fill = hfill
                    c.border = thin
                    col += 1
                gfill = PatternFill(
                    start_color="D1FAE5", end_color="D1FAE5", fill_type="solid"
                )
                rfill = PatternFill(
                    start_color="FEE2E2", end_color="FEE2E2", fill_type="solid"
                )
                afill = PatternFill(
                    start_color="FEF3C7", end_color="FEF3C7", fill_type="solid"
                )
                for ri, s in enumerate(students, 2):
                    ws.cell(ri, 1, s.full_name).border = thin
                    pc = 0
                    for ci, d in enumerate(dates):
                        st = att_map.get(s.id, {}).get(d.isoformat())
                        cell = ws.cell(ri, ci + 2)
                        if st == "present":
                            cell.value = "K"
                            cell.fill = gfill
                            pc += 1
                        elif st == "absent":
                            cell.value = "Y"
                            cell.fill = rfill
                        elif st == "excused":
                            cell.value = "S"
                            cell.fill = afill
                        cell.border = thin
                        cell.alignment = Alignment(horizontal="center")
                    ws.cell(ri, len(dates) + 2, pc).border = thin
                    pct = round(pc / len(dates) * 100) if dates else 0
                    ws.cell(ri, len(dates) + 3, f"{pct}%").border = thin
                buf = io.BytesIO()
                wb.save(buf)
                buf.seek(0)
                await update.message.reply_document(
                    document=buf,
                    filename=f"{g.name}_{today.month}_{today.year}.xlsx",
                    caption=f"📚 {g.name} - {MONTH_NAMES_UZ[today.month - 1]} {today.year}",
                )
            except Exception as e:
                await update.message.reply_text(
                    f"{g.name}: Hisobot tayyorlanmadi ({e})"
                )
        await update.message.reply_text("Barcha hisobotlar yuborildi.")
    finally:
        db.close()


async def cmd_reportpdf(update: Update, context: ContextTypes.DEFAULT_TYPE):
    chat_id = update.effective_chat.id
    db: DBSession = SessionLocal()
    try:
        user = get_user_by_chat_id(db, chat_id)
        if not user:
            await update.message.reply_text("Avval /start orqali hisobingizni ulang.")
            return
        from calendar import monthrange
        from fpdf import FPDF

        today = date.today()
        year, month = today.year, today.month
        if context.args:
            try:
                parts = context.args[0].split("-")
                if len(parts) == 2:
                    month, year = int(parts[1]), int(parts[0])
                elif len(parts) == 1:
                    month = int(parts[0])
            except ValueError:
                await update.message.reply_text(
                    "❌ Noto'g'ri sana. Namuna: /reportpdf 2026-07 yoki /reportpdf 7"
                )
                return
            if not 1 <= month <= 12 or year < 2000:
                await update.message.reply_text("❌ Sana diapazondan tashqarida.")
                return
        await update.message.reply_text("PDF hisobot tayyorlanmoqda, biroz kuting...")
        groups = (
            db.query(models.Group)
            .filter(
                models.Group.user_id == user.id,
                models.Group.is_archived == False,
            )
            .all()
        )
        for g in groups:
            try:
                dates = get_group_dates_for_month(g.days_of_week, year, month)
                students = (
                    db.query(models.Student)
                    .filter(models.Student.group_id == g.id)
                    .order_by(models.Student.id)
                    .all()
                )
                attendances = (
                    db.query(models.Attendance)
                    .filter(
                        models.Attendance.group_id == g.id,
                        models.Attendance.date >= date(year, month, 1),
                        models.Attendance.date
                        <= date(year, month, monthrange(year, month)[1]),
                    )
                    .all()
                )
                att_map = {}
                for a in attendances:
                    att_map.setdefault(a.student_id, {})[a.date.isoformat()] = a.status
                pdf = FPDF(orientation="L", unit="mm", format="A4")
                pdf.add_page()
                pdf.set_font("Helvetica", size=7)
                pdf.set_fill_color(5, 150, 105)
                pdf.set_text_color(255, 255, 255)
                pdf.cell(50, 7, "O'quvchi", border=1, fill=True)
                for d in dates:
                    pdf.cell(12, 7, d.strftime("%d.%m"), border=1, fill=True, align="C")
                pdf.cell(15, 7, "Kelgan", border=1, fill=True, align="C")
                pdf.cell(12, 7, "Foiz", border=1, fill=True, align="C")
                pdf.ln()
                pdf.set_text_color(30, 41, 59)
                for s in students:
                    present = 0
                    pdf.cell(50, 6, s.full_name[:25], border=1)
                    for d in dates:
                        st = att_map.get(s.id, {}).get(d.isoformat())
                        val = {"present": "K", "absent": "Y", "excused": "S"}.get(
                            st, ""
                        )
                        if st == "present":
                            pdf.set_fill_color(209, 250, 229)
                        elif st == "absent":
                            pdf.set_fill_color(254, 226, 226)
                        elif st == "excused":
                            pdf.set_fill_color(254, 243, 199)
                        else:
                            pdf.set_fill_color(255, 255, 255)
                        pdf.cell(12, 6, val, border=1, fill=True, align="C")
                        if st == "present":
                            present += 1
                    pct = round(present / len(dates) * 100) if dates else 0
                    pdf.cell(15, 6, str(present), border=1, align="C")
                    pdf.cell(12, 6, f"{pct}%", border=1, align="C")
                    pdf.ln()
                buf = io.BytesIO(pdf.output())
                buf.seek(0)
                await update.message.reply_document(
                    document=buf,
                    filename=f"{g.name}_{month}_{year}.pdf",
                    caption=f"📚 {g.name} - {MONTH_NAMES_UZ[month - 1]} {year}",
                )
            except Exception as e:
                await update.message.reply_text(f"{g.name}: PDF tayyorlanmadi ({e})")
        await update.message.reply_text("Barcha PDF hisobotlar yuborildi.")
    finally:
        db.close()


async def cmd_payments(update: Update, context: ContextTypes.DEFAULT_TYPE):
    chat_id = update.effective_chat.id
    db: DBSession = SessionLocal()
    try:
        user = get_user_by_chat_id(db, chat_id)
        if not user:
            await update.message.reply_text("Avval /start orqali hisobingizni ulang.")
            return
        today = date.today()
        groups = (
            db.query(models.Group)
            .filter(
                models.Group.user_id == user.id,
                models.Group.is_archived == False,
            )
            .all()
        )
        total_income = 0
        lines = [
            f"<b>To'lov hisoboti - {MONTH_NAMES_UZ[today.month - 1]} {today.year}</b>\n"
        ]
        for g in groups:
            payments = (
                db.query(models.Payment)
                .filter(
                    models.Payment.group_id == g.id,
                    models.Payment.date >= date(today.year, today.month, 1),
                    models.Payment.date
                    <= date(
                        today.year,
                        today.month,
                        __import__("calendar").monthrange(today.year, today.month)[1],
                    ),
                )
                .all()
            )
            if payments:
                group_sum = sum(p.amount for p in payments)
                total_income += group_sum
                lines.append(
                    f"\n📚 <b>{g.name}</b>: {group_sum:,} so'm ({len(payments)} ta)"
                )
        lines.append(f"\n\n💰 <b>Jami daromad: {total_income:,} so'm</b>")
        await update.message.reply_text("\n".join(lines), parse_mode=ParseMode.HTML)
    finally:
        db.close()


async def cmd_schedule(update: Update, context: ContextTypes.DEFAULT_TYPE):
    chat_id = update.effective_chat.id
    db: DBSession = SessionLocal()
    try:
        user = get_user_by_chat_id(db, chat_id)
        if not user:
            links = (
                db.query(models.ParentLink)
                .filter(models.ParentLink.chat_id == str(chat_id))
                .all()
            )
            students = [
                db.query(models.Student)
                .filter(models.Student.id == l.student_id)
                .first()
                for l in links
            ]
            students = [s for s in students if s]
            if not students:
                await update.message.reply_text(
                    "Avval /phone orqali farzandingizga ulaning."
                )
                return
            lines = ["<b>Farzandingiz dars jadvali</b>\n"]
            for s in students:
                g = db.query(models.Group).filter(models.Group.id == s.group_id).first()
                if not g:
                    continue
                days = ", ".join(g.days_of_week or [])
                time = g.lesson_time.strftime("%H:%M") if g.lesson_time else "--:--"
                lines.append(
                    f"👤 <b>{s.full_name}</b>\n   📚 {g.name}\n   🕐 {time} | 📅 {days}"
                )
            await update.message.reply_text("\n".join(lines), parse_mode=ParseMode.HTML)
            return
        groups = (
            db.query(models.Group)
            .filter(models.Group.user_id == user.id, models.Group.is_archived == False)
            .all()
        )
        lines = ["<b>Dars jadvali</b>\n"]
        for g in groups:
            days = ", ".join(g.days_of_week or [])
            time = g.lesson_time.strftime("%H:%M") if g.lesson_time else "--:--"
            sc = (
                db.query(models.Student)
                .filter(
                    models.Student.group_id == g.id, models.Student.status == "active"
                )
                .count()
            )
            lines.append(f"📚 <b>{g.name}</b>\n   🕐 {time} | 📅 {days} | 👥 {sc} ta")
        await update.message.reply_text("\n".join(lines), parse_mode=ParseMode.HTML)
    finally:
        db.close()


async def cmd_topic(update: Update, context: ContextTypes.DEFAULT_TYPE):
    chat_id = update.effective_chat.id
    args = context.args or []
    db: DBSession = SessionLocal()
    try:
        user = get_user_by_chat_id(db, chat_id)
        if not user:
            await update.message.reply_text("Avval /start orqali hisobingizni ulang.")
            return
        today = date.today()
        if args:
            try:
                group_id = int(args[0])
            except ValueError:
                await update.message.reply_text("❌ Guruh raqami noto'g'ri.")
                return
            group = (
                db.query(models.Group)
                .filter(models.Group.id == group_id, models.Group.user_id == user.id)
                .first()
            )
            if not group:
                await update.message.reply_text("❌ Guruh topilmadi.")
                return
            homework = None
            topic_parts = []
            rest = args[1:]
            i = 0
            while i < len(rest):
                if rest[i] in ("--homework", "--hw") and i + 1 < len(rest):
                    homework = rest[i + 1].strip('"')
                    i += 2
                else:
                    topic_parts.append(rest[i])
                    i += 1
            if not topic_parts:
                await update.message.reply_text(
                    "Ishlatish: /topic guruh_raqami Mavzu [--homework 'uy vazifasi']\n"
                    "Masalan: /topic 1 Algebra --homework '4-mashq'"
                )
                return
            topic_text = " ".join(topic_parts)
            topic = (
                db.query(models.LessonTopic)
                .filter(
                    models.LessonTopic.group_id == group_id,
                    models.LessonTopic.date == today,
                )
                .first()
            )
            if topic:
                topic.topic = topic_text
                topic.homework = homework
            else:
                db.add(
                    models.LessonTopic(
                        group_id=group_id,
                        date=today,
                        topic=topic_text,
                        homework=homework,
                    )
                )
            db.commit()
            await update.message.reply_text(
                f"✅ <b>{group.name}</b> mavzusi saqlandi:\n"
                f"   📖 {topic_text}\n"
                f"   📝 {homework or '—'}",
                parse_mode=ParseMode.HTML,
            )
            await send_topic_to_parents(
                db, group_id=group_id, topic_text=topic_text, homework=homework
            )
            return
        groups = (
            db.query(models.Group)
            .filter(models.Group.user_id == user.id, models.Group.is_archived == False)
            .all()
        )
        lines = [f"<b>Bugungi darslar - {today.day}/{today.month}</b>\n"]
        found = False
        for g in groups:
            exc = (
                db.query(models.ExceptionDay)
                .filter(
                    models.ExceptionDay.group_id == g.id,
                    models.ExceptionDay.date == today,
                )
                .first()
            )
            if exc:
                lines.append(f"📚 {g.name} - <b>ISTI SN</b> ({exc.reason or ''})")
                found = True
                continue
            topic = (
                db.query(models.LessonTopic)
                .filter(
                    models.LessonTopic.group_id == g.id,
                    models.LessonTopic.date == today,
                )
                .first()
            )
            if topic and topic.homework:
                lines.append(
                    f"📚 <b>{g.name}</b>: {topic.topic}\n"
                    f"   📝 Uy vazifa: {topic.homework}"
                )
            else:
                lines.append(
                    f"📚 <b>{g.name}</b>: {topic.topic if topic else 'Mavzu belgilanmagan'}"
                )
            found = True
        if not found:
            lines.append("Bugun darslar yo'q.")
        lines.append("\nMavzu qo'shish: /topic guruh_raqami Mavzu --homework 'vazifa'")
        await update.message.reply_text("\n".join(lines), parse_mode=ParseMode.HTML)
    finally:
        db.close()


async def cmd_mark(update: Update, context: ContextTypes.DEFAULT_TYPE):
    chat_id = update.effective_chat.id
    args = context.args or []
    if len(args) < 3:
        await update.message.reply_text(
            "Ishlatish: /mark guruh_raqami o'quvchi_raqami holat\n"
            "Holat: present | absent | excused | none\n\n"
            "Masalan: /mark 1 5 absent\n"
            "Barcha o'quvchilarni belgilash: /mark 1 all present"
        )
        return
    db: DBSession = SessionLocal()
    try:
        user = get_user_by_chat_id(db, chat_id)
        if not user:
            await update.message.reply_text("Avval /start orqali hisobingizni ulang.")
            return
        try:
            group_id = int(args[0])
        except ValueError:
            await update.message.reply_text("❌ Guruh raqami noto'g'ri.")
            return
        group = (
            db.query(models.Group)
            .filter(models.Group.id == group_id, models.Group.user_id == user.id)
            .first()
        )
        if not group:
            await update.message.reply_text("❌ Guruh topilmadi.")
            return
        status = args[2].lower()
        if status not in ("present", "absent", "excused", "none"):
            await update.message.reply_text(
                "❌ Holat noto'g'ri: present | absent | excused | none"
            )
            return
        today = date.today()
        if args[1] == "all":
            students = (
                db.query(models.Student)
                .filter(
                    models.Student.group_id == group_id,
                    models.Student.status == "active",
                )
                .all()
            )
            count = 0
            for s in students:
                att = (
                    db.query(models.Attendance)
                    .filter(
                        models.Attendance.group_id == group_id,
                        models.Attendance.student_id == s.id,
                        models.Attendance.date == today,
                    )
                    .first()
                )
                if status == "none":
                    if att:
                        db.delete(att)
                    continue
                if att:
                    att.status = status
                else:
                    db.add(
                        models.Attendance(
                            group_id=group_id,
                            student_id=s.id,
                            date=today,
                            status=status,
                        )
                    )
                count += 1
            db.commit()
            if status == "absent" and _bot_app is not None:
                for s in students:
                    send_absent_notification(
                        group_name=group.name,
                        student_name=s.full_name,
                        student_id=s.id,
                        date_str=today.isoformat(),
                        teacher_chat_id=user.telegram_chat_id,
                    )
            await update.message.reply_text(
                f"✅ {count} ta o'quvchi '{status}' qilindi."
            )
            return
        try:
            student_id = int(args[1])
        except ValueError:
            await update.message.reply_text("❌ O'quvchi raqami noto'g'ri.")
            return
        student = (
            db.query(models.Student)
            .filter(
                models.Student.id == student_id, models.Student.group_id == group_id
            )
            .first()
        )
        if not student:
            await update.message.reply_text("❌ O'quvchi topilmadi.")
            return
        att = (
            db.query(models.Attendance)
            .filter(
                models.Attendance.group_id == group_id,
                models.Attendance.student_id == student_id,
                models.Attendance.date == today,
            )
            .first()
        )
        if status == "none":
            if att:
                db.delete(att)
                db.commit()
            await update.message.reply_text(
                f"🗑️ {student.full_name} davomati o'chirildi."
            )
            return
        if att:
            att.status = status
        else:
            db.add(
                models.Attendance(
                    group_id=group_id, student_id=student_id, date=today, status=status
                )
            )
        db.commit()
        if status == "absent":
            send_absent_notification(
                group_name=group.name,
                student_name=student.full_name,
                student_id=student.id,
                date_str=today.isoformat(),
                teacher_chat_id=user.telegram_chat_id,
            )
        emoji = {"present": "✅", "absent": "❌", "excused": "⏳"}.get(status, "❓")
        await update.message.reply_text(
            f"{emoji} {student.full_name} — {status} ({today.day}/{today.month})"
        )
    finally:
        db.close()


async def cmd_addstudent(update: Update, context: ContextTypes.DEFAULT_TYPE):
    chat_id = update.effective_chat.id
    args = context.args or []
    if len(args) < 2:
        await update.message.reply_text(
            "Ishlatish: /addstudent guruh_raqami Ism Familiya\n"
            "Masalan: /addstudent 1 Alisher Karimov\n"
            'Qo\'shimcha: /addstudent 1 Alisher Karimov --tel 998901234567 --ota "Ota ismi" --otatel 998901234568 --tugilgan 2015-05-12'
        )
        return
    db: DBSession = SessionLocal()
    try:
        user = get_user_by_chat_id(db, chat_id)
        if not user:
            await update.message.reply_text("Avval /start orqali hisobingizni ulang.")
            return
        try:
            group_id = int(args[0])
        except ValueError:
            await update.message.reply_text("❌ Guruh raqami noto'g'ri.")
            return
        group = (
            db.query(models.Group)
            .filter(models.Group.id == group_id, models.Group.user_id == user.id)
            .first()
        )
        if not group:
            await update.message.reply_text(
                "❌ Guruh topilmadi. /groups orqali raqamni oling."
            )
            return
        rest = args[1:]
        phone = None
        parent_name = None
        parent_phone = None
        birth_date = None
        name_parts = []
        i = 0
        while i < len(rest):
            if rest[i] == "--tel" and i + 1 < len(rest):
                phone = rest[i + 1]
                i += 2
            elif rest[i] == "--ota" and i + 1 < len(rest):
                parent_name = rest[i + 1].strip('"')
                i += 2
            elif rest[i] == "--otatel" and i + 1 < len(rest):
                parent_phone = rest[i + 1]
                i += 2
            elif rest[i] == "--tugilgan" and i + 1 < len(rest):
                try:
                    birth_date = date.fromisoformat(rest[i + 1])
                except ValueError:
                    await update.message.reply_text(
                        "❌ Tug'ilgan sana noto'g'ri. Format: YYYY-MM-DD"
                    )
                    return
                i += 2
            else:
                name_parts.append(rest[i])
                i += 1
        if not name_parts:
            await update.message.reply_text("❌ Ism kiritilmadi.")
            return
        name = " ".join(name_parts)
        student = models.Student(
            group_id=group_id,
            full_name=name,
            phone=phone,
            parent_name=parent_name,
            parent_phone=parent_phone,
            birth_date=birth_date,
            status="active",
        )
        db.add(student)
        db.commit()
        await update.message.reply_text(
            f"✅ <b>{name}</b> {group.name} guruhiga qo'shildi!\n"
            f"   📞 {phone or '—'} | 👪 {parent_name or '—'}",
            parse_mode=ParseMode.HTML,
        )
    finally:
        db.close()


async def cmd_addgroup(update: Update, context: ContextTypes.DEFAULT_TYPE):
    chat_id = update.effective_chat.id
    args = context.args or []
    if len(args) < 2:
        await update.message.reply_text(
            "Ishlatish: /addgroup Nomi KUNLAR Vaqt\n"
            "Kunlar: MON, TUE, WED, THU, FRI, SAT, SUN\n\n"
            "Masalan: /addgroup Matematika MON,WED,FRI 09:00"
        )
        return
    db: DBSession = SessionLocal()
    try:
        user = get_user_by_chat_id(db, chat_id)
        if not user:
            await update.message.reply_text("Avval /start orqali hisobingizni ulang.")
            return
        name = args[0]
        days = [d.upper() for d in args[1].split(",") if d.strip()]
        valid_days = {"MON", "TUE", "WED", "THU", "FRI", "SAT", "SUN"}
        days = [d for d in days if d in valid_days]
        if not days:
            await update.message.reply_text("❌ Kunlar noto'g'ri. Masalan: MON,WED,FRI")
            return
        lesson_time = None
        if len(args) > 2:
            try:
                from datetime import time

                parts = args[2].split(":")
                lesson_time = time(int(parts[0]), int(parts[1]))
            except Exception:
                await update.message.reply_text("❌ Vaqt noto'g'ri. Masalan: 09:00")
                return
        group = models.Group(
            user_id=user.id,
            name=name,
            days_of_week=days,
            lesson_time=lesson_time,
            is_archived=False,
        )
        db.add(group)
        db.commit()
        await update.message.reply_text(
            f"✅ <b>{name}</b> guruhi yaratildi!\n"
            f"   📅 {', '.join(days)}\n"
            f"   🕐 {lesson_time.strftime('%H:%M') if lesson_time else '--:--'}",
            parse_mode=ParseMode.HTML,
        )
    finally:
        db.close()


async def cmd_chart(update: Update, context: ContextTypes.DEFAULT_TYPE):
    chat_id = update.effective_chat.id
    args = context.args or []
    db: DBSession = SessionLocal()
    try:
        user = get_user_by_chat_id(db, chat_id)
        if not user:
            await update.message.reply_text("Avval /start orqali hisobingizni ulang.")
            return
        import matplotlib

        matplotlib.use("Agg")
        import matplotlib.pyplot as plt
        import io

        today = date.today()
        if args:
            try:
                student_id = int(args[0])
            except ValueError:
                await update.message.reply_text("❌ O'quvchi raqami noto'g'ri.")
                return
            student = (
                db.query(models.Student).filter(models.Student.id == student_id).first()
            )
            if not student:
                await update.message.reply_text("❌ O'quvchi topilmadi.")
                return
            group = (
                db.query(models.Group)
                .filter(
                    models.Group.id == student.group_id, models.Group.user_id == user.id
                )
                .first()
            )
            if not group:
                await update.message.reply_text(
                    "❌ Bu o'quvchi sizning guruhlaringizda emas."
                )
                return
            months = []
            pcts = []
            from calendar import monthrange

            for i in range(5, -1, -1):
                y, m = today.year, today.month - i
                while m < 1:
                    m += 12
                    y -= 1
                dates = get_group_dates_for_month(group.days_of_week, y, m)
                exc_days = {
                    e.date
                    for e in (
                        db.query(models.ExceptionDay)
                        .filter(
                            models.ExceptionDay.group_id == group.id,
                            models.ExceptionDay.date >= date(y, m, 1),
                            models.ExceptionDay.date <= date(y, m, monthrange(y, m)[1]),
                        )
                        .all()
                    )
                }
                dates = [d for d in dates if d not in exc_days]
                atts = (
                    db.query(models.Attendance)
                    .filter(
                        models.Attendance.student_id == student.id,
                        models.Attendance.date >= date(y, m, 1),
                        models.Attendance.date <= date(y, m, monthrange(y, m)[1]),
                    )
                    .all()
                )
                att_map = {a.date: a.status for a in atts}
                present = sum(1 for d in dates if att_map.get(d) == "present")
                total = sum(1 for d in dates if att_map.get(d))
                months.append(MONTH_NAMES_UZ[m - 1][:3])
                pcts.append(round(present / total * 100) if total else 0)
            fig, ax = plt.subplots(figsize=(8, 4.2))
            colors = [
                "#059669" if p >= 80 else "#f59e0b" if p >= 50 else "#ef4444"
                for p in pcts
            ]
            bars = ax.bar(
                range(6),
                pcts,
                color=colors,
                width=0.55,
                edgecolor="white",
                linewidth=1.5,
            )
            for i, p in enumerate(pcts):
                ax.text(
                    i,
                    p + 1.5,
                    f"{p}%",
                    ha="center",
                    fontsize=10,
                    fontweight="bold",
                    color="#334155",
                )
            ax.set_ylim(0, 105)
            ax.set_xticks(range(6))
            ax.set_xticklabels(months, fontsize=9)
            ax.set_ylabel("Davomat %", fontsize=9)
            ax.set_title(
                f"{student.full_name} — davomat", fontsize=11, fontweight="bold"
            )
            ax.spines["top"].set_visible(False)
            ax.spines["right"].set_visible(False)
            fig.tight_layout()
            buf = io.BytesIO()
            fig.savefig(buf, format="png", dpi=110)
            buf.seek(0)
            plt.close(fig)
            await update.message.reply_photo(
                photo=buf, caption=f"📊 {student.full_name} — oxirgi 6 oy"
            )
            return
        groups = (
            db.query(models.Group)
            .filter(models.Group.user_id == user.id, models.Group.is_archived == False)
            .all()
        )
        plot_data = []
        for g in groups:
            dates = get_group_dates_for_month(g.days_of_week, today.year, today.month)
            students = (
                db.query(models.Student)
                .filter(
                    models.Student.group_id == g.id, models.Student.status == "active"
                )
                .all()
            )
            if not students or not dates:
                continue
            present = absent = 0
            for d in dates:
                exc = (
                    db.query(models.ExceptionDay)
                    .filter(
                        models.ExceptionDay.group_id == g.id,
                        models.ExceptionDay.date == d,
                    )
                    .first()
                )
                if exc:
                    continue
                for s in students:
                    att = (
                        db.query(models.Attendance)
                        .filter(
                            models.Attendance.student_id == s.id,
                            models.Attendance.date == d,
                        )
                        .first()
                    )
                    if att:
                        if att.status == "present":
                            present += 1
                        elif att.status == "absent":
                            absent += 1
            total = present + absent
            if total > 0:
                plot_data.append((g.name, round(present / total * 100, 1), total))
        if not plot_data:
            await update.message.reply_text("Statistika uchun ma'lumot yetarli emas.")
            return
        fig, ax = plt.subplots(figsize=(8, 4.5))
        names = [p[0][:20] for p in plot_data]
        pcts = [p[1] for p in plot_data]
        colors = [
            "#059669" if p >= 80 else "#f59e0b" if p >= 50 else "#ef4444" for p in pcts
        ]
        bars = ax.bar(
            range(len(plot_data)),
            pcts,
            color=colors,
            width=0.55,
            edgecolor="white",
            linewidth=1.5,
        )
        for i, p in enumerate(plot_data):
            ax.text(
                i,
                pcts[i] + 1.5,
                f"{pcts[i]}%",
                ha="center",
                fontsize=10,
                fontweight="bold",
                color="#334155",
            )
        ax.set_ylim(0, 105)
        ax.set_xticks(range(len(plot_data)))
        ax.set_xticklabels(names, fontsize=8, rotation=12, ha="right")
        ax.set_ylabel("Davomat %", fontsize=9)
        ax.set_title(
            f"Davomat statistikasi - {MONTH_NAMES_UZ[today.month - 1]} {today.year}",
            fontsize=11,
            fontweight="bold",
        )
        ax.spines["top"].set_visible(False)
        ax.spines["right"].set_visible(False)
        fig.tight_layout()
        buf = io.BytesIO()
        fig.savefig(buf, format="png", dpi=110)
        buf.seek(0)
        plt.close(fig)
        await update.message.reply_photo(
            photo=buf, caption=f"📊 {MONTH_NAMES_UZ[today.month - 1]} oyi statistikasi"
        )
    finally:
        db.close()


async def cmd_lang(update: Update, context: ContextTypes.DEFAULT_TYPE):
    chat_id = update.effective_chat.id
    args = context.args or []
    db: DBSession = SessionLocal()
    try:
        user = get_user_by_chat_id(db, chat_id)
        if not args or args[0].lower() not in ("uz", "ru", "en"):
            await update.message.reply_text(
                "Tilni tanlang:\n/lang uz - O'zbekcha\n/lang ru - Русский\n/lang en - English"
            )
            return
        lang = args[0].lower()
        if user:
            user.language = lang
        else:
            link = (
                db.query(models.ParentLink)
                .filter(models.ParentLink.chat_id == str(chat_id))
                .first()
            )
            if link:
                link.language = lang
            else:
                await update.message.reply_text(
                    "Avval /start orqali hisobingizni ulang."
                )
                return
        db.commit()
        msg = {
            "uz": "✅ Til o'zgartirildi: O'zbekcha",
            "ru": "✅ Язык изменён: Русский",
            "en": "✅ Language changed: English",
        }
        await update.message.reply_text(msg[lang])
    finally:
        db.close()


async def cmd_admin(update: Update, context: ContextTypes.DEFAULT_TYPE):
    chat_id = update.effective_chat.id
    args = context.args or []
    db: DBSession = SessionLocal()
    try:
        user = get_user_by_chat_id(db, chat_id)
        if not user or not user.is_superuser:
            await update.message.reply_text("❌ Bu buyruq faqat adminlar uchun.")
            return
        if args and args[0].lower() in ("users", "foydalanuvchilar"):
            users = db.query(models.User).order_by(models.User.created_at.desc()).all()
            lines = [f"<b>Foydalanuvchilar ({len(users)})</b>\n"]
            for u in users:
                gc = db.query(models.Group).filter(models.Group.user_id == u.id).count()
                status = "🟢" if u.is_active else "🔴"
                admin = "👑" if u.is_superuser else ""
                lines.append(
                    f"{status} {admin} <b>{u.full_name}</b>\n   📧 {u.email} | 📚 {gc} guruh | TG: {'✅' if u.telegram_chat_id else '❌'}"
                )
            await update.message.reply_text("\n".join(lines), parse_mode=ParseMode.HTML)
        elif args and args[0].lower() in ("stats", "statistika"):
            users = db.query(models.User).count()
            groups = (
                db.query(models.Group).filter(models.Group.is_archived == False).count()
            )
            students = (
                db.query(models.Student)
                .filter(models.Student.status == "active")
                .count()
            )
            today = date.today()
            att_count = (
                db.query(models.Attendance)
                .filter(models.Attendance.date == today)
                .count()
            )
            payments_total = db.query(models.Payment).count()
            await update.message.reply_text(
                f"📊 <b>Umumiy statistika</b>\n\n"
                f"👥 Foydalanuvchilar: <b>{users}</b>\n"
                f"📚 Guruhlar: <b>{groups}</b>\n"
                f"🎓 O'quvchilar: <b>{students}</b>\n"
                f"📝 Bugungi davomat: <b>{att_count}</b>\n"
                f"💰 Jami to'lovlar: <b>{payments_total}</b> ta",
                parse_mode=ParseMode.HTML,
            )
        else:
            await update.message.reply_text(
                "👑 <b>Admin buyruqlari</b>\n\n"
                "/admin users - Foydalanuvchilar ro'yxati\n"
                "/admin stats - Umumiy statistika\n"
                "/broadcast Xabar - Barchaga xabar yuborish",
                parse_mode=ParseMode.HTML,
            )
    finally:
        db.close()


async def cmd_paymentlist(update: Update, context: ContextTypes.DEFAULT_TYPE):
    chat_id = update.effective_chat.id
    db: DBSession = SessionLocal()
    try:
        user = get_user_by_chat_id(db, chat_id)
        if not user:
            await update.message.reply_text("Avval /start orqali hisobingizni ulang.")
            return
        today = date.today()
        groups = (
            db.query(models.Group)
            .filter(models.Group.user_id == user.id, models.Group.is_archived == False)
            .all()
        )
        lines = [
            f"<b>To'lamaganlar - {MONTH_NAMES_UZ[today.month - 1]} {today.year}</b>\n"
        ]
        any_data = False
        for g in groups:
            students = (
                db.query(models.Student)
                .filter(
                    models.Student.group_id == g.id,
                    models.Student.status == "active",
                )
                .all()
            )
            unpaid = []
            for s in students:
                paid = (
                    db.query(models.Payment)
                    .filter(
                        models.Payment.student_id == s.id,
                        models.Payment.date >= date(today.year, today.month, 1),
                        models.Payment.date
                        <= date(
                            today.year,
                            today.month,
                            __import__("calendar").monthrange(today.year, today.month)[
                                1
                            ],
                        ),
                    )
                    .first()
                )
                if not paid:
                    unpaid.append(s)
            if unpaid:
                any_data = True
                lines.append(f"\n📚 <b>{g.name}</b> ({len(unpaid)} ta):")
                for s in unpaid:
                    lines.append(f"   • {s.full_name}")
        if not any_data:
            lines.append("Hamma to'lab bo'lgan. 🎉")
        await update.message.reply_text("\n".join(lines), parse_mode=ParseMode.HTML)
    finally:
        db.close()


async def cmd_rating(update: Update, context: ContextTypes.DEFAULT_TYPE):
    chat_id = update.effective_chat.id
    args = context.args or []
    db: DBSession = SessionLocal()
    try:
        user = get_user_by_chat_id(db, chat_id)
        if not user:
            await update.message.reply_text("Avval /start orqali hisobingizni ulang.")
            return
        today = date.today()
        if args:
            try:
                group_id = int(args[0])
            except ValueError:
                await update.message.reply_text("❌ Guruh raqami noto'g'ri.")
                return
            groups = (
                db.query(models.Group)
                .filter(models.Group.id == group_id, models.Group.user_id == user.id)
                .all()
            )
        else:
            groups = (
                db.query(models.Group)
                .filter(
                    models.Group.user_id == user.id, models.Group.is_archived == False
                )
                .all()
            )
        if not groups:
            await update.message.reply_text("❌ Guruh topilmadi.")
            return
        output = []
        medals = ["🥇", "🥈", "🥉"]
        for g in groups:
            dates = get_group_dates_for_month(g.days_of_week, today.year, today.month)
            exc_days = {
                e.date
                for e in (
                    db.query(models.ExceptionDay)
                    .filter(
                        models.ExceptionDay.group_id == g.id,
                        models.ExceptionDay.date >= date(today.year, today.month, 1),
                    )
                    .all()
                )
            }
            dates = [d for d in dates if d not in exc_days]
            students = (
                db.query(models.Student)
                .filter(
                    models.Student.group_id == g.id,
                    models.Student.status == "active",
                )
                .all()
            )
            rows = []
            for s in students:
                atts = (
                    db.query(models.Attendance)
                    .filter(
                        models.Attendance.student_id == s.id,
                        models.Attendance.date >= date(today.year, today.month, 1),
                        models.Attendance.date
                        <= date(
                            today.year,
                            today.month,
                            __import__("calendar").monthrange(today.year, today.month)[
                                1
                            ],
                        ),
                    )
                    .all()
                )
                att_map = {a.date: a.status for a in atts}
                present = sum(1 for d in dates if att_map.get(d) == "present")
                total = sum(1 for d in dates if att_map.get(d))
                pct = round(present / total * 100) if total else 0
                rows.append((s.full_name, pct, present, total))
            rows.sort(key=lambda r: r[1], reverse=True)
            lines = [f"\n📚 <b>{g.name}</b> — reyting:\n"]
            for i, (name, pct, present, total) in enumerate(rows):
                medal = medals[i] if i < 3 else f"{i + 1}."
                lines.append(f"{medal} {name} — {pct}% ({present}/{total})")
            if not rows:
                lines.append("Ma'lumot yo'q.")
            output.append("\n".join(lines))
        await update.message.reply_text("\n".join(output), parse_mode=ParseMode.HTML)
    finally:
        db.close()


async def cmd_payment_history(update: Update, context: ContextTypes.DEFAULT_TYPE):
    chat_id = update.effective_chat.id
    args = context.args or []
    db: DBSession = SessionLocal()
    try:
        user = get_user_by_chat_id(db, chat_id)
        today = date.today()
        if user:
            if len(args) < 2:
                await update.message.reply_text(
                    "Ishlatish: /payment-history guruh_raqami o'quvchi_raqami"
                )
                return
            try:
                group_id, student_id = int(args[0]), int(args[1])
            except ValueError:
                await update.message.reply_text("❌ Raqam noto'g'ri.")
                return
            group = (
                db.query(models.Group)
                .filter(models.Group.id == group_id, models.Group.user_id == user.id)
                .first()
            )
            if not group:
                await update.message.reply_text("❌ Guruh topilmadi.")
                return
            student = (
                db.query(models.Student)
                .filter(
                    models.Student.id == student_id, models.Student.group_id == group_id
                )
                .first()
            )
            if not student:
                await update.message.reply_text("❌ O'quvchi topilmadi.")
                return
            students = [student]
        else:
            links = (
                db.query(models.ParentLink)
                .filter(models.ParentLink.chat_id == str(chat_id))
                .all()
            )
            students = [
                db.query(models.Student)
                .filter(models.Student.id == l.student_id)
                .first()
                for l in links
            ]
            students = [s for s in students if s]
            if not students:
                await update.message.reply_text(
                    "Avval /phone orqali farzandingizga ulaning."
                )
                return
        for s in students:
            payments = (
                db.query(models.Payment)
                .filter(models.Payment.student_id == s.id)
                .order_by(models.Payment.date.desc())
                .limit(10)
                .all()
            )
            group = db.query(models.Group).filter(models.Group.id == s.group_id).first()
            lines = [f"💳 <b>{s.full_name}</b> ({group.name if group else '—'})\n"]
            if not payments:
                lines.append("To'lovlar yo'q.")
            for p in payments:
                lines.append(f"   • {p.date.isoformat()} — {p.amount:,} so'm")
            await update.message.reply_text("\n".join(lines), parse_mode=ParseMode.HTML)
    finally:
        db.close()


async def cmd_setcard(update: Update, context: ContextTypes.DEFAULT_TYPE):
    chat_id = update.effective_chat.id
    db: DBSession = SessionLocal()
    try:
        user = get_user_by_chat_id(db, chat_id)
        if not user:
            await update.message.reply_text("Avval /start orqali hisobingizni ulang.")
            return
        card = " ".join(context.args or []).strip()
        if not card:
            await update.message.reply_text(
                "Ishlatish: /setcard 8600123456789012\n"
                "Karta raqami o'rnatilgach, ota-onalar /pay orqali ko'radi."
            )
            return
        user.card_number = card
        db.commit()
        await update.message.reply_text(
            f"✅ Karta raqami o'rnatildi:\n<code>{card}</code>",
            parse_mode=ParseMode.HTML,
        )
    finally:
        db.close()


async def cmd_broadcast(update: Update, context: ContextTypes.DEFAULT_TYPE):
    chat_id = update.effective_chat.id
    args = context.args or []
    db: DBSession = SessionLocal()
    try:
        user = get_user_by_chat_id(db, chat_id)
        if not user or not user.is_superuser:
            await update.message.reply_text("❌ Bu buyruq faqat adminlar uchun.")
            return
        text = " ".join(args)
        if not text:
            await update.message.reply_text(
                "Ishlatish: /broadcast Xabar matni\n"
                "Barcha foydalanuvchilar va ota-onalarga yuboriladi."
            )
            return
        targets = set()
        for u in (
            db.query(models.User).filter(models.User.telegram_chat_id.isnot(None)).all()
        ):
            if u.telegram_chat_id:
                targets.add(int(u.telegram_chat_id))
        for pl in db.query(models.ParentLink).all():
            try:
                targets.add(int(pl.chat_id))
            except ValueError:
                pass
        sent, failed = 0, 0
        for t in targets:
            try:
                await _bot_app.bot.send_message(
                    chat_id=t,
                    text=f"📢 <b>E'lon</b>\n\n{text}",
                    parse_mode=ParseMode.HTML,
                )
                sent += 1
            except Exception:
                failed += 1
            await asyncio.sleep(0.05)
        await update.message.reply_text(
            f"📢 Xabar yuborildi: {sent} ta (muvaffaqiyatsiz: {failed})"
        )
    finally:
        db.close()


async def cmd_editstudent(update: Update, context: ContextTypes.DEFAULT_TYPE):
    chat_id = update.effective_chat.id
    args = context.args or []
    if len(args) < 2:
        await update.message.reply_text(
            "Ishlatish: /editstudent guruh_raqami o'quvchi_raqami [o'zgarishlar]\n"
            'O\'zgarishlar: --ism "Yangi ism" --tel 998901234567 --ota "Ota ismi" '
            "--otatel 998901234568 --tugilgan 2015-05-12 --holat active|inactive\n"
            "Masalan: /editstudent 1 5 --holat inactive"
        )
        return
    db: DBSession = SessionLocal()
    try:
        user = get_user_by_chat_id(db, chat_id)
        if not user:
            await update.message.reply_text("Avval /start orqali hisobingizni ulang.")
            return
        try:
            group_id, student_id = int(args[0]), int(args[1])
        except ValueError:
            await update.message.reply_text("❌ Raqam noto'g'ri.")
            return
        group = (
            db.query(models.Group)
            .filter(models.Group.id == group_id, models.Group.user_id == user.id)
            .first()
        )
        if not group:
            await update.message.reply_text("❌ Guruh topilmadi.")
            return
        student = (
            db.query(models.Student)
            .filter(
                models.Student.id == student_id, models.Student.group_id == group_id
            )
            .first()
        )
        if not student:
            await update.message.reply_text("❌ O'quvchi topilmadi.")
            return
        rest = args[2:]
        changes = []
        i = 0
        while i < len(rest):
            if rest[i] == "--ism" and i + 1 < len(rest):
                student.full_name = rest[i + 1].strip('"')
                changes.append("ism")
                i += 2
            elif rest[i] == "--tel" and i + 1 < len(rest):
                student.phone = rest[i + 1]
                changes.append("telefon")
                i += 2
            elif rest[i] == "--ota" and i + 1 < len(rest):
                student.parent_name = rest[i + 1].strip('"')
                changes.append("ota-ona")
                i += 2
            elif rest[i] == "--otatel" and i + 1 < len(rest):
                student.parent_phone = rest[i + 1]
                changes.append("ota-ona tel")
                i += 2
            elif rest[i] == "--tugilgan" and i + 1 < len(rest):
                try:
                    student.birth_date = date.fromisoformat(rest[i + 1])
                    changes.append("tug'ilgan kun")
                except ValueError:
                    await update.message.reply_text(
                        "❌ Tug'ilgan sana noto'g'ri. Format: YYYY-MM-DD"
                    )
                    return
                i += 2
            elif rest[i] == "--holat" and i + 1 < len(rest):
                st = rest[i + 1].lower()
                if st not in ("active", "inactive"):
                    await update.message.reply_text("❌ Holat: active yoki inactive")
                    return
                student.status = st
                changes.append("holat")
                i += 2
            else:
                await update.message.reply_text(
                    f"❌ Noto'g'ri parametr: {rest[i]}\n"
                    "Parametrlar: --ism --tel --ota --otatel --tugilgan --holat"
                )
                return
        if not changes:
            await update.message.reply_text("❌ Hech qanday o'zgarish kiritilmadi.")
            return
        db.commit()
        await update.message.reply_text(
            f"✅ <b>{student.full_name}</b> yangilandi:\n"
            f"   O'zgargan: {', '.join(changes)}",
            parse_mode=ParseMode.HTML,
        )
    finally:
        db.close()


async def cmd_removestudent(update: Update, context: ContextTypes.DEFAULT_TYPE):
    chat_id = update.effective_chat.id
    args = context.args or []
    if len(args) < 2:
        await update.message.reply_text(
            "Ishlatish: /removestudent guruh_raqami o'quvchi_raqami"
        )
        return
    db: DBSession = SessionLocal()
    try:
        user = get_user_by_chat_id(db, chat_id)
        if not user:
            await update.message.reply_text("Avval /start orqali hisobingizni ulang.")
            return
        try:
            group_id, student_id = int(args[0]), int(args[1])
        except ValueError:
            await update.message.reply_text("❌ Raqam noto'g'ri.")
            return
        group = (
            db.query(models.Group)
            .filter(models.Group.id == group_id, models.Group.user_id == user.id)
            .first()
        )
        if not group:
            await update.message.reply_text("❌ Guruh topilmadi.")
            return
        student = (
            db.query(models.Student)
            .filter(
                models.Student.id == student_id, models.Student.group_id == group_id
            )
            .first()
        )
        if not student:
            await update.message.reply_text("❌ O'quvchi topilmadi.")
            return
        name = student.full_name
        db.delete(student)
        db.commit()
        await update.message.reply_text(f"🗑️ <b>{name}</b> o'chirildi.")
    finally:
        db.close()


async def cmd_archive(update: Update, context: ContextTypes.DEFAULT_TYPE):
    chat_id = update.effective_chat.id
    args = context.args or []
    db: DBSession = SessionLocal()
    try:
        user = get_user_by_chat_id(db, chat_id)
        if not user:
            await update.message.reply_text("Avval /start orqali hisobingizni ulang.")
            return
        if not args:
            groups = (
                db.query(models.Group)
                .filter(models.Group.user_id == user.id)
                .order_by(models.Group.id)
                .all()
            )
            lines = ["<b>Guruhlar (arxivlanganlar * bilan)</b>\n"]
            for g in groups:
                mark = "*" if g.is_archived else ""
                lines.append(
                    f"{g.id}. {g.name}{mark} — /archive {g.id} bilan o'zgartiriladi"
                )
            await update.message.reply_text("\n".join(lines), parse_mode=ParseMode.HTML)
            return
        try:
            group_id = int(args[0])
        except ValueError:
            await update.message.reply_text("❌ Guruh raqami noto'g'ri.")
            return
        group = (
            db.query(models.Group)
            .filter(models.Group.id == group_id, models.Group.user_id == user.id)
            .first()
        )
        if not group:
            await update.message.reply_text("❌ Guruh topilmadi.")
            return
        group.is_archived = not group.is_archived
        db.commit()
        state = "arxivlandi" if group.is_archived else "qaytarildi"
        await update.message.reply_text(f"📚 <b>{group.name}</b> {state}.")
    finally:
        db.close()


async def cmd_setfee(update: Update, context: ContextTypes.DEFAULT_TYPE):
    chat_id = update.effective_chat.id
    args = context.args or []
    db: DBSession = SessionLocal()
    try:
        user = get_user_by_chat_id(db, chat_id)
        if not user:
            await update.message.reply_text("Avval /start orqali hisobingizni ulang.")
            return
        if not args:
            await update.message.reply_text(
                f"Joriy oylik to'lov: {user.monthly_fee or MONTHLY_FEE_UZS:,} so'm\n\n"
                "Ishlatish: /setfee 600000\n"
                "0 kiritsangiz, umumiy narx ishlatiladi."
            )
            return
        try:
            fee = int(args[0].replace(" ", ""))
        except ValueError:
            await update.message.reply_text("❌ Miqdor noto'g'ri.")
            return
        if fee < 0:
            await update.message.reply_text("❌ Miqdor manfiy bo'lishi mumkin emas.")
            return
        user.monthly_fee = fee
        db.commit()
        await update.message.reply_text(
            f"✅ Oylik to'lov: {fee or MONTHLY_FEE_UZS:,} so'm qilib belgilandi."
        )
    finally:
        db.close()


async def cmd_topstudents(update: Update, context: ContextTypes.DEFAULT_TYPE):
    chat_id = update.effective_chat.id
    args = context.args or []
    db: DBSession = SessionLocal()
    try:
        user = get_user_by_chat_id(db, chat_id)
        if not user:
            await update.message.reply_text("Avval /start orqali hisobingizni ulang.")
            return
        today = date.today()
        if args:
            try:
                group_id = int(args[0])
            except ValueError:
                await update.message.reply_text("❌ Guruh raqami noto'g'ri.")
                return
            groups = (
                db.query(models.Group)
                .filter(models.Group.id == group_id, models.Group.user_id == user.id)
                .all()
            )
        else:
            groups = (
                db.query(models.Group)
                .filter(
                    models.Group.user_id == user.id, models.Group.is_archived == False
                )
                .all()
            )
        from calendar import monthrange

        output = []
        medals = ["🥇", "🥈", "🥉"]
        for g in groups:
            dates = get_group_dates_for_month(g.days_of_week, today.year, today.month)
            exc_days = {
                e.date
                for e in (
                    db.query(models.ExceptionDay)
                    .filter(
                        models.ExceptionDay.group_id == g.id,
                        models.ExceptionDay.date >= date(today.year, today.month, 1),
                    )
                    .all()
                )
            }
            dates = [d for d in dates if d not in exc_days]
            students = (
                db.query(models.Student)
                .filter(
                    models.Student.group_id == g.id,
                    models.Student.status == "active",
                )
                .all()
            )
            rows = []
            for s in students:
                atts = (
                    db.query(models.Attendance)
                    .filter(
                        models.Attendance.student_id == s.id,
                        models.Attendance.date >= date(today.year, today.month, 1),
                        models.Attendance.date
                        <= date(
                            today.year,
                            today.month,
                            monthrange(today.year, today.month)[1],
                        ),
                    )
                    .all()
                )
                att_map = {a.date: a.status for a in atts}
                present = sum(1 for d in dates if att_map.get(d) == "present")
                total = sum(1 for d in dates if att_map.get(d))
                pct = round(present / total * 100) if total else 0
                rows.append((s.full_name, pct, present, total))
            rows.sort(key=lambda r: r[1], reverse=True)
            lines = [f"\n📚 <b>{g.name}</b> — oyning eng yaxshilari:\n"]
            for i, (name, pct, present, total) in enumerate(rows[:5]):
                medal = medals[i] if i < 3 else f"{i + 1}."
                lines.append(f"{medal} {name} — {pct}% ({present}/{total})")
            if not rows:
                lines.append("Ma'lumot yo'q.")
            output.append("\n".join(lines))
        await update.message.reply_text("\n".join(output), parse_mode=ParseMode.HTML)
    finally:
        db.close()


async def send_topic_to_parents(
    db: DBSession, group_id: int, topic_text: str, homework: Optional[str] = None
):
    if not _bot_app:
        return
    group = db.query(models.Group).filter(models.Group.id == group_id).first()
    if not group:
        return
    students = (
        db.query(models.Student)
        .filter(models.Student.group_id == group_id, models.Student.status == "active")
        .all()
    )
    student_ids = [s.id for s in students]
    if not student_ids:
        return
    links = (
        db.query(models.ParentLink)
        .filter(models.ParentLink.student_id.in_(student_ids))
        .all()
    )
    hw_text = homework if homework else "yo'q"
    text = (
        f"📚 <b>{group.name}</b> — bugungi dars\n\n"
        f"📖 Mavzu: <b>{topic_text}</b>\n"
        f"📝 Uy vazifa: {hw_text}"
    )
    sent_chats = set()
    for pl in links:
        if pl.chat_id in sent_chats:
            continue
        sent_chats.add(pl.chat_id)
        try:
            await _bot_app.bot.send_message(
                chat_id=int(pl.chat_id), text=text, parse_mode=ParseMode.HTML
            )
        except Exception as e:
            logger.error(f"Failed to send topic to parent: {e}")


PARENT_TEXTS = {
    "absent": {
        "uz": (
            "⚠️ <b>Farzandingiz darsga kelmadi!</b>\n\n"
            "👤 {name}\n📚 {group}\n📅 {date}\n\n"
            "Sababli deb belgilamoqchi bo'lsangiz, pastdagi tugmani bosing."
        ),
        "ru": (
            "⚠️ <b>Ваш ребёнок не пришёл на урок!</b>\n\n"
            "👤 {name}\n📚 {group}\n📅 {date}\n\n"
            "Если это по уважительной причине, нажмите кнопку ниже."
        ),
        "en": (
            "⚠️ <b>Your child missed the lesson!</b>\n\n"
            "👤 {name}\n📚 {group}\n📅 {date}\n\n"
            "If it was for a valid reason, press the button below."
        ),
    },
    "excuse_ok": {
        "uz": "✅ <b>Belgilandi!</b> {name} sababli deb qayd etildi. O'qituvchi xabardor qilindi.",
        "ru": "✅ <b>Отмечено!</b> {name} отмечен как отсутствующий по уважительной причине. Учитель уведомлён.",
        "en": "✅ <b>Done!</b> {name} was marked as excused. The teacher has been notified.",
    },
    "excuse_teacher": {
        "uz": "📝 <b>O'quvchi sababli deb belgilandi</b>\n\n👤 {name}\n📚 {group}\n📅 {date}\n(ota-ona orqali)",
        "ru": "📝 <b>Ученик отмечен как отсутствующий по уважительной причине</b>\n\n👤 {name}\n📚 {group}\n📅 {date}\n(родителем)",
        "en": "📝 <b>Student marked as excused</b>\n\n👤 {name}\n📚 {group}\n📅 {date}\n(by parent)",
    },
    "birthday_student": {
        "uz": "🎂 <b>Bugun tug'ilgan kun!</b>\n\n👤 {name}\n📚 {group}\n🎉 Tabriklaymiz!",
        "ru": "🎂 <b>Сегодня день рождения!</b>\n\n👤 {name}\n📚 {group}\n🎉 Поздравляем!",
        "en": "🎂 <b>Birthday today!</b>\n\n👤 {name}\n📚 {group}\n🎉 Congratulations!",
    },
    "weekly": {
        "uz": (
            "📊 <b>Haftalik xulosa ({week})</b>\n\n"
            "👤 {name}\n📚 {group}\n\n"
            "✅ Kelgan: {present}\n"
            "❌ Yo'q: {absent}\n"
            "⏳ Sababli: {excused}"
        ),
        "ru": (
            "📊 <b>Итоги недели ({week})</b>\n\n"
            "👤 {name}\n📚 {group}\n\n"
            "✅ Присутствовал: {present}\n"
            "❌ Отсутствовал: {absent}\n"
            "⏳ По уважительной причине: {excused}"
        ),
        "en": (
            "📊 <b>Weekly summary ({week})</b>\n\n"
            "👤 {name}\n📚 {group}\n\n"
            "✅ Present: {present}\n"
            "❌ Absent: {absent}\n"
            "⏳ Excused: {excused}"
        ),
    },
}


def _parent_lang(db: DBSession, chat_id: str) -> str:
    link = (
        db.query(models.ParentLink).filter(models.ParentLink.chat_id == chat_id).first()
    )
    if link and link.language:
        return link.language
    try:
        user = get_user_by_chat_id(db, int(chat_id))
    except ValueError:
        user = None
    return user.language if user and user.language else "uz"


def _format_text(db: DBSession, key: str, chat_id: str, **kwargs) -> str:
    lang = _parent_lang(db, chat_id)
    tpl = PARENT_TEXTS[key].get(lang, PARENT_TEXTS[key]["uz"])
    return tpl.format(**kwargs)


async def cmd_excuse_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    if not query:
        return
    await query.answer()
    try:
        _, student_id_str, date_str = query.data.split(":")
        student_id, exc_date = int(student_id_str), date.fromisoformat(date_str)
    except Exception:
        await query.message.reply_text("❌ Xatolik yuz berdi.")
        return
    db: DBSession = SessionLocal()
    try:
        student = (
            db.query(models.Student).filter(models.Student.id == student_id).first()
        )
        if not student:
            await query.message.reply_text("❌ O'quvchi topilmadi.")
            return
        group = (
            db.query(models.Group).filter(models.Group.id == student.group_id).first()
        )
        att = (
            db.query(models.Attendance)
            .filter(
                models.Attendance.student_id == student_id,
                models.Attendance.date == exc_date,
            )
            .first()
        )
        if att:
            att.status = "excused"
        else:
            db.add(
                models.Attendance(
                    group_id=student.group_id,
                    student_id=student_id,
                    date=exc_date,
                    status="excused",
                )
            )
        db.commit()
        try:
            await query.edit_message_text(
                _format_text(
                    db, "excuse_ok", str(query.message.chat_id), name=student.full_name
                ),
                parse_mode=ParseMode.HTML,
            )
        except Exception:
            pass
        user = db.query(models.User).filter(models.User.id == group.user_id).first()
        if user and user.telegram_chat_id:
            try:
                await _bot_app.bot.send_message(
                    chat_id=int(user.telegram_chat_id),
                    text=_format_text(
                        db,
                        "excuse_teacher",
                        user.telegram_chat_id,
                        name=student.full_name,
                        group=group.name,
                        date=exc_date.isoformat(),
                    ),
                    parse_mode=ParseMode.HTML,
                )
            except Exception as e:
                logger.error(f"Failed to notify teacher of excuse: {e}")
    finally:
        db.close()


async def send_absent_notification_async(
    group_name: str,
    student_name: str,
    student_phone: Optional[str],
    date_str: str,
    teacher_chat_id: Optional[str],
):
    if not _bot_app:
        return
    if teacher_chat_id:
        text = (
            f"⚠️ <b>O'quvchi kelmadi!</b>\n\n"
            f"📚 {group_name}\n"
            f"👤 {student_name}\n"
            f"📅 {date_str}\n"
        )
        if student_phone:
            text += f"📞 {student_phone}"
        try:
            await _bot_app.bot.send_message(
                chat_id=int(teacher_chat_id),
                text=text,
                parse_mode=ParseMode.HTML,
            )
        except Exception as e:
            logger.error(f"Failed to send absent notification: {e}")


def notify_receipt_status(receipt_id: int, status: str):
    if not _bot_app:
        return

    def _run(coro):
        try:
            asyncio.run(coro)
        except RuntimeError:
            loop = asyncio.new_event_loop()
            try:
                loop.run_until_complete(coro)
            finally:
                loop.close()

    db: DBSession = SessionLocal()
    try:
        receipt = (
            db.query(models.PaymentReceipt)
            .filter(models.PaymentReceipt.id == receipt_id)
            .first()
        )
        if not receipt:
            return
        student = (
            db.query(models.Student)
            .filter(models.Student.id == receipt.student_id)
            .first()
        )
        group = (
            db.query(models.Group).filter(models.Group.id == receipt.group_id).first()
        )
        if status == "confirmed" and receipt.amount:
            text = (
                f"✅ <b>To'lov tasdiqlandi!</b>\n\n"
                f"👤 {student.full_name if student else '—'}\n"
                f"📚 {group.name if group else '—'}\n"
                f"💵 {receipt.amount:,} so'm\n"
                f"🆔 Chek #{receipt.id}\n\n"
                f"Rahmat! 🙂"
            )
        else:
            text = (
                f"❌ <b>Chek rad etildi</b>\n\n"
                f"👤 {student.full_name if student else '—'}\n"
                f"🆔 Chek #{receipt.id}\n\n"
                f"O'qituvchi bilan bog'lanib, chekni qayta tekshiring."
            )
        if receipt.chat_id:
            try:
                _run(
                    _bot_app.bot.send_message(
                        chat_id=int(receipt.chat_id),
                        text=text,
                        parse_mode=ParseMode.HTML,
                    )
                )
            except Exception as e:
                logger.error(f"Failed to notify parent: {e}")
    finally:
        db.close()


def send_absent_notification(
    group_name: str,
    student_name: str,
    student_id: int,
    date_str: str,
    teacher_chat_id: Optional[str],
):
    db: DBSession = SessionLocal()
    try:
        student = (
            db.query(models.Student).filter(models.Student.id == student_id).first()
        )
        student_phone = student.phone if student else None
        parent_links = (
            db.query(models.ParentLink)
            .filter(models.ParentLink.student_id == student_id)
            .all()
        )
    finally:
        db.close()

    def _run(coro):
        try:
            asyncio.run(coro)
        except RuntimeError:
            loop = asyncio.new_event_loop()
            try:
                loop.run_until_complete(coro)
            finally:
                loop.close()

    _run(
        send_absent_notification_async(
            group_name=group_name,
            student_name=student_name,
            student_phone=student_phone,
            date_str=date_str,
            teacher_chat_id=teacher_chat_id,
        )
    )

    if parent_links:
        for pl in parent_links:
            db2: DBSession = SessionLocal()
            try:
                parent_text = _format_text(
                    db2,
                    "absent",
                    pl.chat_id,
                    name=student_name,
                    group=group_name,
                    date=date_str,
                )
            finally:
                db2.close()
            keyboard = InlineKeyboardMarkup(
                [
                    [
                        InlineKeyboardButton(
                            "✅ Sababli qilish",
                            callback_data=f"excuse:{student_id}:{date_str}",
                        )
                    ]
                ]
            )
            try:
                _run(
                    _bot_app.bot.send_message(
                        chat_id=int(pl.chat_id),
                        text=parent_text,
                        parse_mode=ParseMode.HTML,
                        reply_markup=keyboard,
                    )
                )
            except Exception as e:
                logger.error(f"Failed to send parent notification: {e}")


async def send_daily_report():
    db: DBSession = SessionLocal()
    try:
        today = date.today()
        weekday_name = ["MON", "TUE", "WED", "THU", "FRI", "SAT", "SUN"][
            today.weekday()
        ]
        users = (
            db.query(models.User)
            .filter(
                models.User.telegram_chat_id.isnot(None),
                models.User.is_active == True,
            )
            .all()
        )
        for user in users:
            groups = (
                db.query(models.Group)
                .filter(
                    models.Group.user_id == user.id,
                    models.Group.is_archived == False,
                )
                .all()
            )
            for g in groups:
                if weekday_name not in (g.days_of_week or []):
                    continue
                exc = (
                    db.query(models.ExceptionDay)
                    .filter(
                        models.ExceptionDay.group_id == g.id,
                        models.ExceptionDay.date == today,
                    )
                    .first()
                )
                if exc:
                    continue
                students = (
                    db.query(models.Student)
                    .filter(
                        models.Student.group_id == g.id,
                        models.Student.status == "active",
                    )
                    .all()
                )
                if not students:
                    continue
                present = absent = excused = 0
                absent_list = []
                for s in students:
                    att = (
                        db.query(models.Attendance)
                        .filter(
                            models.Attendance.group_id == g.id,
                            models.Attendance.student_id == s.id,
                            models.Attendance.date == today,
                        )
                        .first()
                    )
                    if att:
                        if att.status == "present":
                            present += 1
                        elif att.status == "absent":
                            absent += 1
                            absent_list.append(s.full_name)
                        elif att.status == "excused":
                            excused += 1
                total = len(students)
                text = (
                    f"📊 <b>Kunlik hisobot</b>\n"
                    f"📚 {g.name}\n"
                    f"📅 {today.day}/{today.month}/{today.year}\n\n"
                    f"✅ Kelgan: {present}\n"
                    f"❌ Yo'q: {absent}\n"
                    f"⏳ Sababli: {excused}\n"
                    f"📊 {total} ta o'quvchi\n"
                )
                if absent_list:
                    text += f"\n<b>Kelmaganlar:</b>\n" + "\n".join(
                        f"• {n}" for n in absent_list
                    )
                try:
                    await _bot_app.bot.send_message(
                        chat_id=int(user.telegram_chat_id),
                        text=text,
                        parse_mode=ParseMode.HTML,
                    )
                except Exception as e:
                    logger.error(
                        f"Failed to send daily report to {user.telegram_chat_id}: {e}"
                    )
    finally:
        db.close()


async def send_lesson_reminders():
    db: DBSession = SessionLocal()
    try:
        today = date.today()
        weekday_name = ["MON", "TUE", "WED", "THU", "FRI", "SAT", "SUN"][
            today.weekday()
        ]
        now = datetime.now()
        users = (
            db.query(models.User)
            .filter(
                models.User.telegram_chat_id.isnot(None),
                models.User.is_active == True,
            )
            .all()
        )
        for user in users:
            groups = (
                db.query(models.Group)
                .filter(
                    models.Group.user_id == user.id,
                    models.Group.is_archived == False,
                )
                .all()
            )
            for g in groups:
                if weekday_name not in (g.days_of_week or []):
                    continue
                if not g.lesson_time:
                    continue
                lt = g.lesson_time
                start_delta = (lt.hour * 60 + lt.minute) - (now.hour * 60 + now.minute)
                if 25 <= start_delta <= 40:
                    try:
                        await _bot_app.bot.send_message(
                            chat_id=int(user.telegram_chat_id),
                            text=(
                                f"⏰ <b>Darsga 30 daqiqa qoldi!</b>\n\n"
                                f"📚 {g.name}\n"
                                f"🕐 {lt.strftime('%H:%M')}"
                            ),
                            parse_mode=ParseMode.HTML,
                        )
                    except Exception as e:
                        logger.error(f"Lesson reminder failed: {e}")
    finally:
        db.close()


async def send_birthday_reminders():
    if not _bot_app:
        return
    db: DBSession = SessionLocal()
    try:
        today = date.today()
        students = (
            db.query(models.Student).filter(models.Student.birth_date.isnot(None)).all()
        )
        for s in students:
            if s.birth_date.month != today.month or s.birth_date.day != today.day:
                continue
            group = db.query(models.Group).filter(models.Group.id == s.group_id).first()
            teacher = (
                db.query(models.User).filter(models.User.id == group.user_id).first()
                if group
                else None
            )
            if teacher and teacher.telegram_chat_id:
                try:
                    await _bot_app.bot.send_message(
                        chat_id=int(teacher.telegram_chat_id),
                        text=_format_text(
                            db,
                            "birthday_student",
                            teacher.telegram_chat_id,
                            name=s.full_name,
                            group=group.name if group else "—",
                        ),
                        parse_mode=ParseMode.HTML,
                    )
                except Exception as e:
                    logger.error(f"Failed to send birthday to teacher: {e}")
            links = (
                db.query(models.ParentLink)
                .filter(models.ParentLink.student_id == s.id)
                .all()
            )
            for pl in links:
                try:
                    await _bot_app.bot.send_message(
                        chat_id=int(pl.chat_id),
                        text=_format_text(
                            db,
                            "birthday_student",
                            pl.chat_id,
                            name=s.full_name,
                            group=group.name if group else "—",
                        ),
                        parse_mode=ParseMode.HTML,
                    )
                except Exception as e:
                    logger.error(f"Failed to send birthday to parent: {e}")
    finally:
        db.close()


async def send_weekly_summary():
    if not _bot_app:
        return
    db: DBSession = SessionLocal()
    try:
        today = date.today()
        monday = today - timedelta(days=today.weekday())
        next_monday = monday + timedelta(days=7)
        users = (
            db.query(models.User)
            .filter(
                models.User.telegram_chat_id.isnot(None),
                models.User.is_active == True,
            )
            .all()
        )
        for user in users:
            groups = (
                db.query(models.Group)
                .filter(
                    models.Group.user_id == user.id,
                    models.Group.is_archived == False,
                )
                .all()
            )
            if not groups:
                continue
            lines = [
                f"📊 <b>Haftalik xulosa</b> ({monday.day}/{monday.month} — {today.day}/{today.month})\n"
            ]
            total_students = 0
            total_att = 0
            total_present = 0
            for g in groups:
                students = (
                    db.query(models.Student)
                    .filter(
                        models.Student.group_id == g.id,
                        models.Student.status == "active",
                    )
                    .all()
                )
                for s in students:
                    atts = (
                        db.query(models.Attendance)
                        .filter(
                            models.Attendance.student_id == s.id,
                            models.Attendance.date >= monday,
                            models.Attendance.date < next_monday,
                        )
                        .all()
                    )
                    for a in atts:
                        total_att += 1
                        if a.status == "present":
                            total_present += 1
                    total_students += 1
                week_pay = (
                    db.query(models.Payment)
                    .filter(
                        models.Payment.group_id == g.id,
                        models.Payment.date >= monday,
                        models.Payment.date < next_monday,
                    )
                    .all()
                )
                income = sum(p.amount for p in week_pay)
                if income:
                    lines.append(f"   💰 {g.name}: +{income:,} so'm")
            pct = round(total_present / total_att * 100) if total_att else 0
            lines.insert(1, f"   ✅ Davomat: {pct}% ({total_present}/{total_att})")
            lines.append(
                f"\nJami: {total_students} ta o'quvchi, {total_att} ta davomat"
            )
            try:
                await _bot_app.bot.send_message(
                    chat_id=int(user.telegram_chat_id),
                    text="\n".join(lines),
                    parse_mode=ParseMode.HTML,
                )
            except Exception as e:
                logger.error(f"Failed to send weekly summary to teacher: {e}")
        links = db.query(models.ParentLink).all()
        for pl in links:
            student = (
                db.query(models.Student)
                .filter(models.Student.id == pl.student_id)
                .first()
            )
            if not student:
                continue
            atts = (
                db.query(models.Attendance)
                .filter(
                    models.Attendance.student_id == student.id,
                    models.Attendance.date >= monday,
                    models.Attendance.date < next_monday,
                )
                .all()
            )
            if not atts:
                continue
            present = sum(1 for a in atts if a.status == "present")
            absent = sum(1 for a in atts if a.status == "absent")
            excused = sum(1 for a in atts if a.status == "excused")
            group = (
                db.query(models.Group)
                .filter(models.Group.id == student.group_id)
                .first()
            )
            text = _format_text(
                db,
                "weekly",
                pl.chat_id,
                name=student.full_name,
                group=group.name if group else "—",
                present=present,
                absent=absent,
                excused=excused,
                week=f"{monday.day}/{monday.month}—{today.day}/{today.month}",
            )
            try:
                await _bot_app.bot.send_message(
                    chat_id=int(pl.chat_id),
                    text=text,
                    parse_mode=ParseMode.HTML,
                )
            except Exception as e:
                logger.error(f"Failed to send weekly summary to parent: {e}")
    finally:
        db.close()


async def send_payment_reminders():
    db: DBSession = SessionLocal()
    try:
        today = date.today()
        last_10_days = today - timedelta(days=10)
        weekday_name = ["MON", "TUE", "WED", "THU", "FRI", "SAT", "SUN"][
            today.weekday()
        ]
        users = (
            db.query(models.User)
            .filter(
                models.User.telegram_chat_id.isnot(None),
                models.User.is_active == True,
            )
            .all()
        )
        for user in users:
            groups = (
                db.query(models.Group)
                .filter(
                    models.Group.user_id == user.id,
                    models.Group.is_archived == False,
                )
                .all()
            )
            for g in groups:
                if weekday_name not in (g.days_of_week or []):
                    continue
                students = (
                    db.query(models.Student)
                    .filter(
                        models.Student.group_id == g.id,
                        models.Student.status == "active",
                    )
                    .all()
                )
                unpaid = []
                for s in students:
                    has_recent_payment = (
                        db.query(models.Payment)
                        .filter(
                            models.Payment.student_id == s.id,
                            models.Payment.date >= last_10_days,
                        )
                        .first()
                    )
                    if not has_recent_payment:
                        unpaid.append(s)
                if unpaid:
                    try:
                        await _bot_app.bot.send_message(
                            chat_id=int(user.telegram_chat_id),
                            text=(
                                f"💰 <b>To'lov eslatmasi</b>\n"
                                f"📚 {g.name}\n\n"
                                f"So'nggi 10 kunda to'lov qilmagan o'quvchilar:\n"
                                + "\n".join(f"• {s.full_name}" for s in unpaid[:10])
                                + (
                                    f"\n… va yana {len(unpaid) - 10} ta"
                                    if len(unpaid) > 10
                                    else ""
                                )
                            ),
                            parse_mode=ParseMode.HTML,
                        )
                    except Exception as e:
                        logger.error(f"Payment reminder failed: {e}")
    finally:
        db.close()


async def scheduler():
    last_pay_day = -1
    last_birthday_day = -1
    last_weekly_day = -1
    while True:
        now = datetime.now()
        if now.hour == 19 and now.minute == 0:
            await send_daily_report()
            await asyncio.sleep(61)
        if now.hour == 20 and now.minute == 0 and now.day != last_pay_day:
            last_pay_day = now.day
            await send_payment_reminders()
            await asyncio.sleep(61)
        if now.hour == 9 and now.minute == 0 and now.day != last_birthday_day:
            last_birthday_day = now.day
            await send_birthday_reminders()
            await asyncio.sleep(61)
        if (
            now.weekday() == 6
            and now.hour == 18
            and now.minute == 0
            and now.day != last_weekly_day
        ):
            last_weekly_day = now.day
            await send_weekly_summary()
            await asyncio.sleep(61)
        if now.minute % 5 == 0:
            await send_lesson_reminders()
            await asyncio.sleep(31)
        await asyncio.sleep(30)


async def post_init(app: Application):
    global _bot_app
    _bot_app = app
    try:
        await app.bot.set_my_commands(
            [BotCommand(command=c, description=d) for c, d in COMMANDS_MENU]
        )
    except Exception as e:
        logger.error(f"set_my_commands failed: {e}")
    asyncio.create_task(scheduler())


KEYBOARD_CMDS = {
    "📋 Bugungi davomat": "today",
    "📊 Statistika": "stats",
    "📚 Guruhlarim": "groups",
    "💰 To'lov hisoboti": "payments",
    "🧾 To'lov qilish": "pay",
    "⭐ Reyting": "rating",
    "📄 Hisobot (Excel)": "report",
    "❓ Yordam": "help",
    "👶 Farzandim": "mystudent",
    "💳 To'lov tarixi": "paymenthistory",
}

_HANDLER_BY_NAME = {}


def _resolve_handler(name: str):
    if _HANDLER_BY_NAME:
        return _HANDLER_BY_NAME.get(name)
    for h in _bot_app.handlers[0]:
        if isinstance(h, CommandHandler):
            for c in h.commands:
                _HANDLER_BY_NAME[c] = h.callback
    return _HANDLER_BY_NAME.get(name)


async def cmd_keyboard_text(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = (update.message.text or "").strip()
    name = KEYBOARD_CMDS.get(text)
    if not name:
        return
    await context.bot.send_chat_action(
        chat_id=update.effective_chat.id, action="typing"
    )
    fn = _resolve_handler(name)
    if fn:
        try:
            await fn(update, context)
        except Exception as e:
            logger.error(f"keyboard command {name} failed: {e}")
            await update.message.reply_text(
                "❌ Xatolik yuz berdi. Qayta urinib ko'ring."
            )


async def cb_receipt_ready(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await q.answer()
    await q.message.reply_text(
        "📸 Yaxshi! Endi chekni (o'tkazma kvitansiyasi) rasm sifatida yuboring.\n"
        "Rasm yuborilgach, o'qituvchingizga xabar boradi va tasdiqlangach sizga bildiriladi."
    )


def build_app():
    app = Application.builder().token(TOKEN).post_init(post_init).build()

    app.add_handler(CommandHandler("start", cmd_start))
    app.add_handler(CommandHandler("link", cmd_link))
    app.add_handler(CommandHandler("help", cmd_help))
    app.add_handler(CommandHandler("groups", cmd_groups))
    app.add_handler(CommandHandler("today", cmd_today))
    app.add_handler(CommandHandler("stats", cmd_stats))
    app.add_handler(CommandHandler("student", cmd_student))
    app.add_handler(CommandHandler("report", cmd_report))
    app.add_handler(CommandHandler("reportpdf", cmd_reportpdf))
    app.add_handler(CommandHandler("payments", cmd_payments))
    app.add_handler(CommandHandler("phone", cmd_phone))
    app.add_handler(CommandHandler("select", cmd_select))
    app.add_handler(CommandHandler("mystudent", cmd_mystudent))
    app.add_handler(CommandHandler("pay", cmd_pay))
    app.add_handler(CommandHandler("schedule", cmd_schedule))
    app.add_handler(CommandHandler("topic", cmd_topic))
    app.add_handler(CommandHandler("mark", cmd_mark))
    app.add_handler(CommandHandler("addstudent", cmd_addstudent))
    app.add_handler(CommandHandler("addgroup", cmd_addgroup))
    app.add_handler(CommandHandler("chart", cmd_chart))
    app.add_handler(CommandHandler("lang", cmd_lang))
    app.add_handler(CommandHandler("admin", cmd_admin))
    app.add_handler(CommandHandler("paymentlist", cmd_paymentlist))
    app.add_handler(CommandHandler("rating", cmd_rating))
    app.add_handler(CommandHandler("paymenthistory", cmd_payment_history))
    app.add_handler(CommandHandler("broadcast", cmd_broadcast))
    app.add_handler(CommandHandler("editstudent", cmd_editstudent))
    app.add_handler(CommandHandler("removestudent", cmd_removestudent))
    app.add_handler(CommandHandler("archive", cmd_archive))
    app.add_handler(CommandHandler("setfee", cmd_setfee))
    app.add_handler(CommandHandler("setcard", cmd_setcard))
    app.add_handler(CommandHandler("topstudents", cmd_topstudents))
    app.add_handler(CallbackQueryHandler(cmd_excuse_callback, pattern=r"^excuse:"))
    app.add_handler(CallbackQueryHandler(cb_receipt_ready, pattern=r"^receipt:ready$"))
    app.add_handler(MessageHandler(filters.PHOTO, cmd_receipt_photo))
    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, cmd_keyboard_text))

    return app


def run_bot():
    loop = asyncio.new_event_loop()
    asyncio.set_event_loop(loop)
    app = build_app()
    try:
        loop.run_until_complete(app.initialize())
        loop.run_until_complete(app.start())
        if WEBHOOK_URL:
            loop.run_until_complete(
                app.bot.set_webhook(
                    url=WEBHOOK_URL,
                    secret_token=WEBHOOK_SECRET,
                    allowed_updates=Update.ALL_TYPES,
                )
            )
            loop.run_until_complete(
                app.updater.start_webhook(
                    listen="0.0.0.0",
                    port=WEBHOOK_PORT,
                    url_path=TOKEN,
                    secret_token=WEBHOOK_SECRET,
                )
            )
            logger.warning(f"Bot webhook started on port {WEBHOOK_PORT}")
        else:
            loop.run_until_complete(
                app.updater.start_polling(
                    allowed_updates=Update.ALL_TYPES, poll_interval=5
                )
            )
        loop.run_forever()
    except (KeyboardInterrupt, SystemExit):
        pass
    except Exception as e:
        logger.warning(f"Bot polling stopped: {e}")
    finally:
        try:
            loop.run_until_complete(app.updater.stop())
            loop.run_until_complete(app.stop())
            loop.run_until_complete(app.shutdown())
        except Exception:
            pass
        loop.close()
